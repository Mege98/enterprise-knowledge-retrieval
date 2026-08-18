#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Windows Keyword Search

A small local indexing/search utility for Windows. It converts supported files
to extracted text, chunks the text, stores it in SQLite, and reports exact
keyword locations as path + line/column/character offset.
"""

from __future__ import annotations

import argparse
import bisect
import ctypes
import csv
import datetime as dt
import html.parser
import io
import json
import os
import queue
import re
import shlex
import sqlite3
import struct
import subprocess
import sys
import threading
import time
import urllib.error
import urllib.request
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Callable, Iterable, Iterator, Optional
from xml.etree import ElementTree as ET


for stream_name in ("stdout", "stderr"):
    stream = getattr(sys, stream_name, None)
    if stream is not None and hasattr(stream, "reconfigure"):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass

DISPLAY_NAME = "企业知识检索系统"
DISPLAY_SUBTITLE = "企业知识检索与可核验问答系统"
AI_REVIEW_HINT = "AI 可能出错，关键信息请人工核实。"
SECURITY_HINT = "本地检索后调用 API 生成答案，AI 可能出错，关键信息请人工核实。请勿输入涉密、敏感、客户隐私以及未公开的经营数据。Enter 发送，Shift+Enter 换行。"
APP_NAME = "EnterpriseKnowledgeRetrieval"
DEFAULT_CHUNK_SIZE = 4000
DEFAULT_CHUNK_OVERLAP = 200
DEFAULT_MAX_FILE_MB = 100
DEFAULT_MAX_TEXT_CHARS = 3_000_000
DEFAULT_RAG_API_URL = "https://api.openai.com/v1/chat/completions"
DEFAULT_RAG_MODEL = "gpt-4o-mini"
DEFAULT_RAG_TOP_K = 6
DEFAULT_RAG_MAX_CONTEXT_CHARS = 9000

TEXT_EXTENSIONS = {
    ".txt",
    ".md",
    ".markdown",
    ".csv",
    ".tsv",
    ".json",
    ".jsonl",
    ".xml",
    ".html",
    ".htm",
    ".log",
    ".ini",
    ".cfg",
    ".conf",
    ".yaml",
    ".yml",
    ".py",
    ".js",
    ".ts",
    ".tsx",
    ".jsx",
    ".java",
    ".c",
    ".cpp",
    ".h",
    ".hpp",
    ".cs",
    ".go",
    ".rs",
    ".php",
    ".rb",
    ".sql",
    ".bat",
    ".cmd",
    ".ps1",
    ".rtf",
}

OFFICE_EXTENSIONS = {".docx", ".xlsx", ".xlsm", ".xls", ".pptx"}
PDF_EXTENSIONS = {".pdf"}
SUPPORTED_EXTENSIONS = TEXT_EXTENSIONS | OFFICE_EXTENSIONS | PDF_EXTENSIONS
RAG_DOCUMENT_EXTENSIONS = OFFICE_EXTENSIONS | PDF_EXTENSIONS | {".txt", ".md", ".markdown", ".rtf"}
RAG_LOW_PRIORITY_EXTENSIONS = TEXT_EXTENSIONS - {".txt", ".md", ".markdown", ".rtf"}


@dataclass
class ConvertedDocument:
    text: str
    converter: str
    warning: str = ""


@dataclass
class SearchHit:
    path: str
    keyword: str
    position: str
    line: int
    column: int
    offset: int
    chunk_no: int
    snippet: str
    occurrence: int
    page: int = 0
    sheet: str = ""
    cell: str = ""


@dataclass
class SourceLocation:
    label: str
    page: int = 0
    sheet: str = ""
    cell: str = ""


@dataclass
class RagContext:
    path: str
    keyword: str
    position: str
    chunk_no: int
    text: str
    score: float = 0.0
    matched_terms: tuple[str, ...] = ()
    origin: str = ""


@dataclass
class FileNameCandidate:
    path: str
    score: float
    matched_terms: tuple[str, ...] = ()


class ExtractHtmlText(html.parser.HTMLParser):
    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: list[str] = []
        self.skip_depth = 0

    def handle_starttag(self, tag: str, attrs: list[tuple[str, Optional[str]]]) -> None:
        if tag.lower() in {"script", "style", "noscript"}:
            self.skip_depth += 1
        if tag.lower() in {"p", "br", "div", "section", "article", "li", "tr", "h1", "h2", "h3", "h4"}:
            self.parts.append("\n")

    def handle_endtag(self, tag: str) -> None:
        if tag.lower() in {"script", "style", "noscript"} and self.skip_depth:
            self.skip_depth -= 1
        if tag.lower() in {"p", "div", "section", "article", "li", "tr"}:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if not self.skip_depth and data.strip():
            self.parts.append(data)

    def text(self) -> str:
        return normalize_text(" ".join(self.parts))


def default_db_path() -> Path:
    base = os.environ.get("LOCALAPPDATA")
    if base:
        root = Path(base) / APP_NAME
    else:
        root = Path.home() / f".{APP_NAME.lower()}"
    return root / "search_index.sqlite"


def default_settings_path() -> Path:
    return default_db_path().with_name("settings.json")


def resource_path(name: str) -> Path:
    base = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
    return base / name


def now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")


def normalize_text(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[ \t\f\v]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def looks_binary(data: bytes) -> bool:
    if not data:
        return False
    if data.startswith((b"\xff\xfe", b"\xfe\xff", b"\xef\xbb\xbf")):
        return False
    if b"\x00" in data[:4096]:
        return True
    text_like = sum(1 for b in data[:4096] if b in b"\n\r\t" or 32 <= b <= 126 or b >= 128)
    return text_like / min(len(data), 4096) < 0.85


def read_text_bytes(path: Path) -> tuple[str, str]:
    raw = path.read_bytes()
    if looks_binary(raw):
        raise ValueError("binary file is not text-decodable")

    encodings = ["utf-8-sig", "utf-16", "gb18030", "cp936", "latin-1"]
    last_error: Optional[Exception] = None
    for encoding in encodings:
        try:
            return raw.decode(encoding), encoding
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"cannot decode text: {last_error}")


def strip_rtf(text: str) -> str:
    text = re.sub(r"\\'[0-9a-fA-F]{2}", " ", text)
    text = re.sub(r"\\[a-zA-Z]+\d* ?", " ", text)
    text = text.replace("{", " ").replace("}", " ")
    return normalize_text(text)


def extract_plain_or_markup(path: Path) -> ConvertedDocument:
    raw_text, encoding = read_text_bytes(path)
    ext = path.suffix.lower()
    warning = ""

    if ext in {".html", ".htm"}:
        parser = ExtractHtmlText()
        parser.feed(raw_text)
        return ConvertedDocument(parser.text(), f"html:{encoding}", warning)

    if ext == ".xml":
        try:
            root = ET.fromstring(raw_text)
            return ConvertedDocument(normalize_text("\n".join(root.itertext())), f"xml:{encoding}", warning)
        except ET.ParseError:
            warning = "XML parse failed; indexed as plain text"

    if ext == ".json":
        try:
            obj = json.loads(raw_text)
            return ConvertedDocument(json.dumps(obj, ensure_ascii=False, indent=2), f"json:{encoding}", warning)
        except json.JSONDecodeError:
            warning = "JSON parse failed; indexed as plain text"

    if ext == ".rtf":
        return ConvertedDocument(strip_rtf(raw_text), f"rtf:{encoding}", warning)

    return ConvertedDocument(normalize_text(raw_text), f"text:{encoding}", warning)


def xml_local_name(tag: str) -> str:
    if "}" in tag:
        return tag.rsplit("}", 1)[1]
    return tag


def paragraph_texts(root: ET.Element, paragraph_tag: str = "p", text_tag: str = "t") -> list[str]:
    lines: list[str] = []
    for elem in root.iter():
        if xml_local_name(elem.tag) == paragraph_tag:
            parts: list[str] = []
            for child in elem.iter():
                if xml_local_name(child.tag) == text_tag and child.text:
                    parts.append(child.text)
            line = "".join(parts).strip()
            if line:
                lines.append(line)
    return lines


def extract_docx(path: Path) -> ConvertedDocument:
    names: list[str] = []
    with zipfile.ZipFile(path) as zf:
        for name in zf.namelist():
            if (
                name == "word/document.xml"
                or name.startswith("word/header")
                or name.startswith("word/footer")
                or name.startswith("word/footnotes")
                or name.startswith("word/endnotes")
            ) and name.endswith(".xml"):
                names.append(name)

        lines: list[str] = []
        for name in names:
            root = ET.fromstring(zf.read(name))
            lines.extend(paragraph_texts(root))

    return ConvertedDocument(normalize_text("\n".join(lines)), "docx")


def extract_pptx(path: Path) -> ConvertedDocument:
    with zipfile.ZipFile(path) as zf:
        slide_names = sorted(
            name for name in zf.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        lines: list[str] = []
        for index, name in enumerate(slide_names, start=1):
            root = ET.fromstring(zf.read(name))
            slide_lines = paragraph_texts(root)
            if slide_lines:
                lines.append(f"[Slide {index}]")
                lines.extend(slide_lines)

    return ConvertedDocument(normalize_text("\n".join(lines)), "pptx")


def get_child_text(elem: ET.Element, child_name: str) -> str:
    for child in elem:
        if xml_local_name(child.tag) == child_name:
            return "".join(child.itertext())
    return ""


def combine_warning(*parts: str) -> str:
    return "; ".join(part for part in parts if part)


def append_limited_line(lines: list[str], total_chars: int, line: str, char_limit: Optional[int]) -> tuple[int, bool]:
    line = str(line).strip()
    if not line:
        return total_chars, False
    separator = 1 if lines else 0
    needed = separator + len(line)
    if char_limit is not None and total_chars + needed > char_limit:
        remaining = max(0, char_limit - total_chars - separator)
        if remaining > 0:
            lines.append(line[:remaining])
        return char_limit or total_chars, True
    lines.append(line)
    return total_chars + needed, False


def spreadsheet_col_name(index: int) -> str:
    result = ""
    while index > 0:
        index, rem = divmod(index - 1, 26)
        result = chr(65 + rem) + result
    return result or "A"


def spreadsheet_cell_ref(row: int, col: int) -> str:
    return f"{spreadsheet_col_name(col)}{row}"


def spreadsheet_value_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def xlsx_sheet_lookup(zf: zipfile.ZipFile) -> dict[str, tuple[int, str]]:
    names = set(zf.namelist())
    workbook_path = "xl/workbook.xml"
    rels_path = "xl/_rels/workbook.xml.rels"
    if workbook_path not in names:
        return {}

    rels: dict[str, str] = {}
    if rels_path in names:
        rel_root = ET.fromstring(zf.read(rels_path))
        for rel in rel_root:
            rel_id = rel.attrib.get("Id", "")
            target = rel.attrib.get("Target", "")
            if not rel_id or not target:
                continue
            if target.startswith("/"):
                normalized = target.lstrip("/")
            elif target.startswith("xl/"):
                normalized = target
            else:
                normalized = f"xl/{target}"
            rels[rel_id] = normalized.replace("\\", "/")

    lookup: dict[str, tuple[int, str]] = {}
    root = ET.fromstring(zf.read(workbook_path))
    index = 1
    for sheet in root.iter():
        if xml_local_name(sheet.tag) != "sheet":
            continue
        name = sheet.attrib.get("name", f"Sheet{index}")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id", "")
        target = rels.get(rel_id)
        if target:
            lookup[target] = (index, name)
        index += 1
    return lookup


def extract_xlsx_zip(path: Path, text_limit: Optional[int] = None) -> ConvertedDocument:
    with zipfile.ZipFile(path) as zf:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in zf.namelist():
            root = ET.fromstring(zf.read("xl/sharedStrings.xml"))
            for si in root.iter():
                if xml_local_name(si.tag) == "si":
                    text = "".join(si.itertext()).strip()
                    shared_strings.append(text)

        sheet_names = sorted(
            name for name in zf.namelist() if name.startswith("xl/worksheets/sheet") and name.endswith(".xml")
        )
        sheet_lookup = xlsx_sheet_lookup(zf)
        lines: list[str] = []
        total_chars = 0
        truncated = False
        for sheet_index, name in enumerate(sheet_names, start=1):
            display_index, display_name = sheet_lookup.get(name, (sheet_index, f"Sheet{sheet_index}"))
            with zf.open(name) as sheet_file:
                for _event, cell in ET.iterparse(sheet_file, events=("end",)):
                    if xml_local_name(cell.tag) != "c":
                        cell.clear()
                        continue
                    cell_ref = cell.attrib.get("r", "")

                    cell_type = cell.attrib.get("t")
                    if cell_type == "s":
                        value_text = get_child_text(cell, "v")
                        try:
                            value = shared_strings[int(value_text)]
                        except (ValueError, IndexError):
                            value = value_text
                    elif cell_type == "inlineStr":
                        value = "".join(cell.itertext()).strip()
                    else:
                        value = get_child_text(cell, "v")
                    if value:
                        cell_label = cell_ref or "Cell"
                        line = f"[Sheet {display_index}: {display_name}] {cell_label}: {value}"
                        total_chars, truncated = append_limited_line(lines, total_chars, line, text_limit)
                        if truncated:
                            break
                    cell.clear()
            if truncated:
                break

    warning = "extracted text truncated at configured limit" if truncated else ""
    return ConvertedDocument(normalize_text("\n".join(lines)), "xlsx:zip", warning)


def extract_xlsx(path: Path, text_limit: Optional[int] = None) -> ConvertedDocument:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception:
        return extract_xlsx_zip(path, text_limit)

    lines: list[str] = []
    total_chars = 0
    truncated = False
    workbook = load_workbook(filename=path, read_only=True, data_only=True, keep_links=False)
    try:
        for sheet_index, sheet in enumerate(workbook.worksheets, start=1):
            sheet_name = sheet.title or f"Sheet{sheet_index}"
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                for col_index, value in enumerate(row, start=1):
                    value_text = spreadsheet_value_text(value)
                    if not value_text:
                        continue
                    cell_label = spreadsheet_cell_ref(row_index, col_index)
                    line = f"[Sheet {sheet_index}: {sheet_name}] {cell_label}: {value_text}"
                    total_chars, truncated = append_limited_line(lines, total_chars, line, text_limit)
                    if truncated:
                        break
                if truncated:
                    break
            if truncated:
                break
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()

    warning = "extracted text truncated at configured limit" if truncated else ""
    return ConvertedDocument(normalize_text("\n".join(lines)), "xlsx:openpyxl", warning)


OLE_FREE_SECTOR = 0xFFFFFFFF
OLE_END_OF_CHAIN = 0xFFFFFFFE
OLE_FAT_SECTOR = 0xFFFFFFFD
OLE_DIFAT_SECTOR = 0xFFFFFFFC


def ole_u16(data: bytes, offset: int) -> int:
    return struct.unpack_from("<H", data, offset)[0]


def ole_u32(data: bytes, offset: int) -> int:
    return struct.unpack_from("<I", data, offset)[0]


def ole_sector(data: bytes, sector_id: int, sector_size: int) -> bytes:
    start = (sector_id + 1) * sector_size
    return data[start : start + sector_size]


def ole_read_chain(data: bytes, fat: list[int], start_sector: int, sector_size: int) -> bytes:
    if start_sector in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN}:
        return b""
    pieces: list[bytes] = []
    seen: set[int] = set()
    sector = start_sector
    while sector not in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN} and sector not in seen:
        if sector < 0 or sector >= len(fat):
            break
        seen.add(sector)
        pieces.append(ole_sector(data, sector, sector_size))
        sector = fat[sector]
    return b"".join(pieces)


def read_ole_stream(data: bytes, stream_names: Iterable[str]) -> Optional[bytes]:
    if len(data) < 512 or data[:8] != b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return None

    sector_size = 1 << ole_u16(data, 30)
    mini_sector_size = 1 << ole_u16(data, 32)
    first_dir_sector = ole_u32(data, 48)
    mini_cutoff = ole_u32(data, 56)
    first_minifat_sector = ole_u32(data, 60)
    first_difat_sector = ole_u32(data, 68)
    num_difat_sectors = ole_u32(data, 72)

    difat = [entry for entry in struct.unpack_from("<109I", data, 76) if entry != OLE_FREE_SECTOR]
    next_difat = first_difat_sector
    for _ in range(num_difat_sectors):
        if next_difat in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN}:
            break
        sector_data = ole_sector(data, next_difat, sector_size)
        entries = list(struct.unpack_from(f"<{sector_size // 4}I", sector_data, 0))
        difat.extend(entry for entry in entries[:-1] if entry != OLE_FREE_SECTOR)
        next_difat = entries[-1]

    fat: list[int] = []
    for sector_id in difat:
        if sector_id in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN, OLE_FAT_SECTOR, OLE_DIFAT_SECTOR}:
            continue
        sector_data = ole_sector(data, sector_id, sector_size)
        fat.extend(struct.unpack_from(f"<{sector_size // 4}I", sector_data, 0))
    if not fat:
        return None

    directory = ole_read_chain(data, fat, first_dir_sector, sector_size)
    entries: dict[str, tuple[int, int, int]] = {}
    root_entry: Optional[tuple[int, int, int]] = None
    for offset in range(0, len(directory) - 127, 128):
        entry = directory[offset : offset + 128]
        name_len = ole_u16(entry, 64)
        entry_type = entry[66]
        if name_len < 2 or entry_type == 0:
            continue
        name = entry[: name_len - 2].decode("utf-16le", errors="ignore")
        start_sector = ole_u32(entry, 116)
        size = struct.unpack_from("<Q", entry, 120)[0]
        item = (entry_type, start_sector, size)
        entries[name.lower()] = item
        if entry_type == 5:
            root_entry = item

    wanted = {name.lower() for name in stream_names}
    target = next((entry for name, entry in entries.items() if name in wanted), None)
    if target is None:
        return None

    _entry_type, start_sector, size = target
    if size < mini_cutoff and root_entry is not None and first_minifat_sector not in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN}:
        _root_type, root_start, root_size = root_entry
        mini_stream = ole_read_chain(data, fat, root_start, sector_size)[:root_size]
        minifat_stream = ole_read_chain(data, fat, first_minifat_sector, sector_size)
        minifat = list(struct.unpack_from(f"<{len(minifat_stream) // 4}I", minifat_stream, 0))
        pieces: list[bytes] = []
        seen: set[int] = set()
        sector = start_sector
        while sector not in {OLE_FREE_SECTOR, OLE_END_OF_CHAIN} and sector not in seen:
            if sector < 0 or sector >= len(minifat):
                break
            seen.add(sector)
            start = sector * mini_sector_size
            pieces.append(mini_stream[start : start + mini_sector_size])
            sector = minifat[sector]
        return b"".join(pieces)[:size]

    return ole_read_chain(data, fat, start_sector, sector_size)[:size]


def iter_biff_records(data: bytes) -> Iterator[tuple[int, int, bytes]]:
    offset = 0
    while offset + 4 <= len(data):
        record_id, size = struct.unpack_from("<HH", data, offset)
        start = offset + 4
        end = start + size
        if end > len(data):
            break
        yield offset, record_id, data[start:end]
        offset = end


def parse_xl_unicode_string(data: bytes, offset: int = 0, short_len: bool = False) -> tuple[str, int]:
    if short_len:
        if offset + 2 > len(data):
            return "", offset
        char_count = data[offset]
        offset += 1
    else:
        if offset + 3 > len(data):
            return "", offset
        char_count = ole_u16(data, offset)
        offset += 2

    options = data[offset]
    offset += 1
    rich_runs = 0
    ext_size = 0
    if options & 0x08:
        if offset + 2 > len(data):
            return "", offset
        rich_runs = ole_u16(data, offset)
        offset += 2
    if options & 0x04:
        if offset + 4 > len(data):
            return "", offset
        ext_size = ole_u32(data, offset)
        offset += 4

    is_utf16 = bool(options & 0x01)
    byte_count = char_count * (2 if is_utf16 else 1)
    raw = data[offset : offset + byte_count]
    if len(raw) < byte_count:
        return "", len(data)
    text = raw.decode("utf-16le" if is_utf16 else "latin-1", errors="ignore").strip()
    offset += byte_count + rich_runs * 4 + ext_size
    return text, offset


def parse_biff_sst(records: list[tuple[int, int, bytes]]) -> list[str]:
    result: list[str] = []
    index = 0
    while index < len(records):
        _offset, record_id, payload = records[index]
        if record_id != 0x00FC:
            index += 1
            continue
        merged = bytearray(payload)
        index += 1
        while index < len(records) and records[index][1] == 0x003C:
            merged.extend(records[index][2])
            index += 1
        data = bytes(merged)
        if len(data) < 8:
            continue
        pos = 8
        while pos < len(data):
            text, new_pos = parse_xl_unicode_string(data, pos)
            if new_pos <= pos:
                break
            if text:
                result.append(text)
            pos = new_pos
    return result


def parse_biff_sheet_name(payload: bytes) -> str:
    if len(payload) < 8:
        return ""
    text, _pos = parse_xl_unicode_string(payload, 6, short_len=True)
    return text


def is_useful_binary_text(text: str) -> bool:
    cleaned = normalize_text(text)
    if len(cleaned) < 2:
        return False
    return any(ch.isalnum() or "\u4e00" <= ch <= "\u9fff" for ch in cleaned)


def extract_raw_binary_strings(data: bytes, text_limit: Optional[int] = None) -> list[str]:
    results: list[str] = []
    seen: set[str] = set()
    total_chars = 0

    def add(text: str) -> bool:
        nonlocal total_chars
        text = normalize_text(text)
        if not is_useful_binary_text(text) or text in seen:
            return False
        seen.add(text)
        total_chars, truncated = append_limited_line(results, total_chars, f"[XLS] {text}", text_limit)
        return truncated

    for start_offset in (0, 1):
        chars: list[str] = []
        for index in range(start_offset, len(data) - 1, 2):
            codepoint = data[index] | (data[index + 1] << 8)
            char = chr(codepoint)
            if char.isprintable() and char not in "\x00\xff":
                chars.append(char)
            else:
                if len(chars) >= 2 and add("".join(chars)):
                    return results
                chars = []
        if len(chars) >= 2 and add("".join(chars)):
            return results

    chunk = bytearray()
    for byte in data:
        if byte in (9, 10, 13) or 32 <= byte <= 126 or byte >= 128:
            chunk.append(byte)
        else:
            if len(chunk) >= 4:
                for encoding in ("gb18030", "latin-1"):
                    try:
                        if add(bytes(chunk).decode(encoding, errors="ignore")):
                            return results
                        break
                    except Exception:
                        continue
            chunk.clear()
    if len(chunk) >= 4:
        for encoding in ("gb18030", "latin-1"):
            try:
                add(bytes(chunk).decode(encoding, errors="ignore"))
                break
            except Exception:
                continue
    return results


def extract_xls(path: Path, text_limit: Optional[int] = None) -> ConvertedDocument:
    raw = path.read_bytes()
    workbook = read_ole_stream(raw, ("Workbook", "Book")) or raw
    records = list(iter_biff_records(workbook))
    shared_strings = parse_biff_sst(records)

    sheet_offsets: dict[int, tuple[int, str]] = {}
    sheet_no = 1
    for _offset, record_id, payload in records:
        if record_id != 0x0085 or len(payload) < 8:
            continue
        sheet_offset = ole_u32(payload, 0)
        name = parse_biff_sheet_name(payload) or f"Sheet{sheet_no}"
        sheet_offsets[sheet_offset] = (sheet_no, name)
        sheet_no += 1

    lines: list[str] = []
    total_chars = 0
    truncated = False
    current_sheet = (1, "Sheet1")

    for offset, record_id, payload in records:
        if offset in sheet_offsets:
            current_sheet = sheet_offsets[offset]

        value = ""
        row = col = 0
        if record_id == 0x0204 and len(payload) >= 9:
            row, col = struct.unpack_from("<HH", payload, 0)
            value, _pos = parse_xl_unicode_string(payload, 6)
        elif record_id == 0x00FD and len(payload) >= 10:
            row, col = struct.unpack_from("<HH", payload, 0)
            sst_index = ole_u32(payload, 6)
            if 0 <= sst_index < len(shared_strings):
                value = shared_strings[sst_index]
        elif record_id == 0x0203 and len(payload) >= 14:
            row, col = struct.unpack_from("<HH", payload, 0)
            number = struct.unpack_from("<d", payload, 6)[0]
            value = spreadsheet_value_text(number)

        if not value:
            continue
        sheet_index, sheet_name = current_sheet
        cell_label = spreadsheet_cell_ref(row + 1, col + 1)
        line = f"[Sheet {sheet_index}: {sheet_name}] {cell_label}: {value}"
        total_chars, truncated = append_limited_line(lines, total_chars, line, text_limit)
        if truncated:
            break

    warning = ""
    converter = "xls:biff"
    if not lines:
        lines = extract_raw_binary_strings(workbook, text_limit)
        converter = "xls:binary-scan"
        warning = "legacy XLS indexed with approximate binary text scan"
    if truncated:
        warning = combine_warning(warning, "extracted text truncated at configured limit")
    return ConvertedDocument(normalize_text("\n".join(lines)), converter, warning)


def extract_pdf(path: Path) -> ConvertedDocument:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return extract_pdf_rough(path)

    reader = PdfReader(str(path))
    lines: list[str] = []
    for index, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            text = f"[Page {index} extraction failed: {exc}]"
        if text.strip():
            lines.append(f"[Page {index}]")
            lines.append(text)
    return ConvertedDocument(normalize_text("\n".join(lines)), "pdf:pypdf")


def extract_pdf_rough(path: Path) -> ConvertedDocument:
    raw = path.read_bytes()
    warning = "Install optional dependency pypdf for better PDF text extraction"
    try:
        decoded = raw.decode("latin-1", errors="ignore")
        pieces = re.findall(r"\((.{1,200}?)\)", decoded, flags=re.DOTALL)
        cleaned = [re.sub(r"\\[nrtbf()\\]", " ", p) for p in pieces]
        return ConvertedDocument(normalize_text("\n".join(cleaned)), "pdf:rough", warning)
    except Exception as exc:
        raise ValueError(f"PDF extraction failed and pypdf is not installed: {exc}") from exc


def convert_file(path: Path, text_limit: Optional[int] = None) -> ConvertedDocument:
    ext = path.suffix.lower()
    if ext == ".docx":
        return extract_docx(path)
    if ext in {".xlsx", ".xlsm"}:
        return extract_xlsx(path, text_limit)
    if ext == ".xls":
        return extract_xls(path, text_limit)
    if ext == ".pptx":
        return extract_pptx(path)
    if ext == ".pdf":
        return extract_pdf(path)
    if ext in TEXT_EXTENSIONS or ext == "":
        return extract_plain_or_markup(path)

    # Unknown extensions are indexed if they are actually text.
    return extract_plain_or_markup(path)


def line_starts_for(text: str) -> list[int]:
    starts = [0]
    for index, char in enumerate(text):
        if char == "\n":
            starts.append(index + 1)
    return starts


def offset_to_line_col(line_starts: list[int], offset: int) -> tuple[int, int]:
    line_index = bisect.bisect_right(line_starts, offset) - 1
    line_start = line_starts[max(0, line_index)]
    return line_index + 1, offset - line_start + 1


def chunk_text(text: str, chunk_size: int, overlap: int) -> Iterator[tuple[int, int, int, str]]:
    if chunk_size < 256:
        chunk_size = 256
    overlap = max(0, min(overlap, chunk_size // 2))
    text_length = len(text)
    line_starts = line_starts_for(text)

    start = 0
    chunk_no = 0
    while start < text_length:
        end = min(start + chunk_size, text_length)
        if end < text_length:
            newline = text.rfind("\n", start + chunk_size // 2, end)
            space = text.rfind(" ", start + chunk_size // 2, end)
            split_at = max(newline, space)
            if split_at > start:
                end = split_at + 1

        chunk = text[start:end]
        line, col = offset_to_line_col(line_starts, start)
        yield chunk_no, start, line, col, chunk

        if end >= text_length:
            break
        start = max(end - overlap, start + 1)
        chunk_no += 1


def init_db(db_path: Path, rebuild: bool = True) -> sqlite3.Connection:
    db_path.parent.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(db_path))
    conn.execute("PRAGMA journal_mode=WAL")
    conn.execute("PRAGMA synchronous=NORMAL")
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS files (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            path TEXT NOT NULL UNIQUE,
            size INTEGER NOT NULL,
            mtime REAL NOT NULL,
            ext TEXT NOT NULL,
            status TEXT NOT NULL,
            converter TEXT,
            warning TEXT,
            error TEXT,
            chars INTEGER NOT NULL DEFAULT 0,
            chunks INTEGER NOT NULL DEFAULT 0,
            indexed_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE IF NOT EXISTS chunks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_id INTEGER NOT NULL,
            chunk_no INTEGER NOT NULL,
            start_offset INTEGER NOT NULL,
            start_line INTEGER NOT NULL,
            start_col INTEGER NOT NULL,
            text TEXT NOT NULL,
            FOREIGN KEY(file_id) REFERENCES files(id) ON DELETE CASCADE
        )
        """
    )
    conn.execute("CREATE INDEX IF NOT EXISTS idx_chunks_file_id ON chunks(file_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_files_status ON files(status)")
    if rebuild:
        conn.execute("DELETE FROM chunks")
        conn.execute("DELETE FROM files")
        conn.commit()
    return conn


def is_probably_skippable(path: Path, root: Path, db_path: Path) -> bool:
    try:
        resolved = path.resolve()
        resolved_db = db_path.resolve()
        db_sidecars = {
            resolved_db,
            Path(str(resolved_db) + "-wal"),
            Path(str(resolved_db) + "-shm"),
            Path(str(resolved_db) + "-journal"),
        }
        if resolved in db_sidecars:
            return True
    except OSError:
        pass
    lowered_parts = {part.lower() for part in path.parts}
    if {"$recycle.bin", "system volume information"} & lowered_parts:
        return True
    try:
        if path.is_dir() and path.resolve() == root.resolve():
            return False
    except OSError:
        pass
    return False


def iter_files(root: Path, db_path: Path) -> Iterator[Path]:
    for current_root, dirs, files in os.walk(root):
        current_path = Path(current_root)
        dirs[:] = [
            d
            for d in dirs
            if not is_probably_skippable(current_path / d, root, db_path)
            and d.lower() not in {".git", "__pycache__", "node_modules"}
        ]
        for name in files:
            path = current_path / name
            if not is_probably_skippable(path, root, db_path):
                yield path


def insert_file_record(
    conn: sqlite3.Connection,
    path: Path,
    status: str,
    converter: str = "",
    warning: str = "",
    error: str = "",
    chars: int = 0,
    chunks: int = 0,
) -> int:
    stat = path.stat()
    cur = conn.execute(
        """
        INSERT OR REPLACE INTO files
            (path, size, mtime, ext, status, converter, warning, error, chars, chunks, indexed_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            str(path),
            stat.st_size,
            stat.st_mtime,
            path.suffix.lower(),
            status,
            converter,
            warning,
            error,
            chars,
            chunks,
            now_iso(),
        ),
    )
    return int(cur.lastrowid)


ProgressCallback = Callable[[dict[str, object]], None]
StopCheck = Callable[[], bool]


def set_process_low_priority(enabled: bool) -> None:
    if os.name != "nt":
        return
    try:
        priority = 0x00004000 if enabled else 0x00000020
        handle = ctypes.windll.kernel32.GetCurrentProcess()
        ctypes.windll.kernel32.SetPriorityClass(handle, priority)
    except Exception:
        pass


def count_files(root: Path, db_path: Path, should_stop: Optional[StopCheck] = None, progress: Optional[ProgressCallback] = None) -> int:
    total = 0
    last_update = 0.0
    for _path in iter_files(root, db_path):
        if should_stop and should_stop():
            break
        total += 1
        now = time.monotonic()
        if progress and now - last_update >= 0.35:
            progress({"event": "counting", "total": total})
            last_update = now
    if progress:
        progress({"event": "counted", "total": total})
    return total


def index_folder(
    folder: Path,
    db_path: Path,
    chunk_size: int = DEFAULT_CHUNK_SIZE,
    overlap: int = DEFAULT_CHUNK_OVERLAP,
    max_file_mb: int = DEFAULT_MAX_FILE_MB,
    max_text_chars: int = DEFAULT_MAX_TEXT_CHARS,
    progress: Optional[ProgressCallback] = None,
    should_stop: Optional[StopCheck] = None,
) -> dict[str, int]:
    folder = folder.resolve()
    db_path = db_path.resolve()
    if not folder.exists() or not folder.is_dir():
        raise ValueError(f"Folder does not exist or is not a directory: {folder}")

    stats = {
        "seen": 0,
        "indexed": 0,
        "skipped": 0,
        "failed": 0,
        "chunks": 0,
        "chars": 0,
        "total": 0,
        "cancelled": 0,
    }
    stats["total"] = count_files(folder, db_path, should_stop=should_stop, progress=progress)
    if should_stop and should_stop():
        stats["cancelled"] = 1
        if progress:
            progress({"event": "cancelled", **stats})
        return stats

    conn = init_db(db_path, rebuild=True)
    max_bytes = max_file_mb * 1024 * 1024
    last_progress = 0.0

    try:
        for path in iter_files(folder, db_path):
            if should_stop and should_stop():
                stats["cancelled"] = 1
                break

            stats["seen"] += 1
            now = time.monotonic()
            if progress and (now - last_progress >= 0.25 or stats["seen"] == stats["total"]):
                progress({"event": "file", "path": str(path), **stats})
                last_progress = now

            try:
                size = path.stat().st_size
                if size > max_bytes:
                    insert_file_record(conn, path, "skipped", error=f"larger than {max_file_mb} MB")
                    stats["skipped"] += 1
                    continue

                ext = path.suffix.lower()
                if ext not in SUPPORTED_EXTENSIONS and ext:
                    insert_file_record(conn, path, "skipped", error="unsupported file type")
                    stats["skipped"] += 1
                    continue
                else:
                    doc = convert_file(path, max_text_chars)

                text = normalize_text(doc.text)
                if not text:
                    insert_file_record(conn, path, "skipped", doc.converter, doc.warning, "no extractable text")
                    stats["skipped"] += 1
                    continue
                if len(text) > max_text_chars:
                    text = text[:max_text_chars]
                    doc.warning = combine_warning(doc.warning, f"extracted text truncated at {max_text_chars} chars")

                file_id = insert_file_record(
                    conn,
                    path,
                    "indexed",
                    converter=doc.converter,
                    warning=doc.warning,
                    chars=len(text),
                    chunks=0,
                )
                batch: list[tuple[int, int, int, int, int, str]] = []
                chunk_count = 0
                for chunk_no, start_offset, start_line, start_col, chunk in chunk_text(text, chunk_size, overlap):
                    batch.append((file_id, chunk_no, start_offset, start_line, start_col, chunk))
                    chunk_count += 1
                    if len(batch) >= 64:
                        conn.executemany(
                            """
                            INSERT INTO chunks
                                (file_id, chunk_no, start_offset, start_line, start_col, text)
                            VALUES (?, ?, ?, ?, ?, ?)
                            """,
                            batch,
                        )
                        batch.clear()
                if batch:
                    conn.executemany(
                        """
                        INSERT INTO chunks
                            (file_id, chunk_no, start_offset, start_line, start_col, text)
                        VALUES (?, ?, ?, ?, ?, ?)
                        """,
                        batch,
                    )
                conn.execute("UPDATE files SET chunks = ? WHERE id = ?", (chunk_count, file_id))
                stats["indexed"] += 1
                stats["chunks"] += chunk_count
                stats["chars"] += len(text)
            except Exception as exc:
                try:
                    insert_file_record(conn, path, "failed", error=str(exc))
                except Exception:
                    pass
                stats["failed"] += 1

            if stats["seen"] % 100 == 0:
                conn.commit()
            time.sleep(0.003)
        conn.commit()
    finally:
        conn.close()

    if progress:
        progress({"event": "cancelled" if stats["cancelled"] else "done", **stats})
    return stats


def parse_keywords(query: str, phrase: bool = False) -> list[str]:
    query = query.strip()
    if not query:
        return []
    if phrase:
        return [query]
    try:
        terms = shlex.split(query)
    except ValueError:
        terms = query.split()
    if not terms:
        terms = [query]
    return [term for term in terms if term]


RAG_QUERY_NOISE_WORDS = {
    "什么",
    "什么是",
    "是什么",
    "请问",
    "请",
    "帮我",
    "帮忙",
    "一下",
    "介绍",
    "解释",
    "说明",
    "定义",
    "含义",
    "哪些",
    "哪个",
    "哪项",
    "怎么",
    "怎样",
    "如何",
    "为什么",
    "是否",
    "有没有",
    "多少",
    "谁",
    "负责",
    "关于",
    "相关",
    "有关",
    "根据",
    "文件",
    "文档",
    "资料",
    "里面",
    "里",
    "中",
    "的",
    "了",
    "吗",
    "呢",
}


RAG_TERM_ALIASES = {
    # 人事行政
    "请假": ("休假", "假期", "年休假", "病假", "事假", "考勤"),
    "休假": ("请假", "假期", "年休假", "调休", "考勤"),
    "员工": ("职工", "人员", "岗位", "入职", "离职"),
    "工资": ("薪资", "薪酬", "报酬", "工资表"),
    "培训": ("学习", "教育", "培养", "考核"),
    # 财务采购
    "报销": ("费用", "票据", "发票", "付款", "审批"),
    "采购": ("供应商", "询价", "合同", "订单", "物资"),
    "付款": ("支付", "结算", "报销", "发票"),
    # 仓储设备质量
    "仓库": ("库存", "入库", "出库", "物料", "盘点"),
    "质量": ("质检", "检验", "不合格", "缺陷", "整改"),
    "工艺": ("作业指导书", "操作规程", "流程", "SOP", "工序"),
    "设备": ("维修", "保养", "点检", "故障", "检修"),
    "安全": ("风险", "隐患", "防护", "应急", "事故"),
    # RAG/检索
    "文件": ("文档", "资料", "制度", "标准", "规程"),
    "规定": ("制度", "要求", "办法", "规范", "标准"),
}


def unique_keep_order(items: Iterable[str]) -> list[str]:
    result: list[str] = []
    seen: set[str] = set()
    for item in items:
        item = normalize_text(item)
        if not item or item in seen:
            continue
        seen.add(item)
        result.append(item)
    return result


def expand_rag_terms(terms: Iterable[str]) -> list[str]:
    """Add a small set of business-domain aliases to improve recall.

    This is not a replacement for embeddings, but it fixes many cases where
    the user asks with one word while the document uses another, e.g.
    “请假” vs “休假/考勤”. Original terms stay first, aliases are appended.
    """
    expanded: list[str] = []
    for term in terms:
        term = normalize_text(term)
        if not term:
            continue
        expanded.append(term)
        for key, aliases in RAG_TERM_ALIASES.items():
            if term == key or key in term or term in key:
                expanded.extend(aliases)
    return unique_keep_order(expanded)


def compact_question_text(text: str) -> str:
    text = normalize_text(text)
    text = re.sub(r"[？?！!，,。；;：:、（）()【】\[\]《》<>\"'“”‘’]+", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    for word in sorted(RAG_QUERY_NOISE_WORDS, key=len, reverse=True):
        text = text.replace(word, " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def rag_search_candidates(question: str) -> list[str]:
    raw = normalize_text(question)
    cleaned = compact_question_text(question)
    candidates: list[str] = []
    if raw:
        candidates.append(raw)
    if cleaned and cleaned != raw:
        candidates.append(cleaned)
    candidates.extend(term for term in parse_keywords(cleaned) if len(term) >= 2)

    joined = re.sub(r"\s+", "", cleaned)
    segments = re.findall(r"[\u4e00-\u9fffA-Za-z0-9]{2,}", joined)
    for segment in segments:
        if len(segment) <= 8:
            candidates.append(segment)
            if re.fullmatch(r"[\u4e00-\u9fff]+", segment) and len(segment) >= 3:
                for length in range(min(5, len(segment)), 1, -1):
                    for start in range(0, len(segment) - length + 1):
                        piece = segment[start : start + length]
                        if piece not in RAG_QUERY_NOISE_WORDS:
                            candidates.append(piece)
            continue
        for length in range(min(8, len(segment)), 1, -1):
            for start in range(0, len(segment) - length + 1):
                piece = segment[start : start + length]
                if piece not in RAG_QUERY_NOISE_WORDS:
                    candidates.append(piece)
            if len(candidates) > 80:
                break
    return unique_keep_order(candidates)[:100]


def rag_question_terms(question: str) -> list[str]:
    raw = normalize_text(question)
    cleaned = compact_question_text(raw)
    terms: list[str] = []
    for source in (cleaned, raw):
        terms.extend(term for term in parse_keywords(source) if len(term) >= 2)
        terms.extend(re.findall(r"[A-Za-z0-9_]{2,}|[\u4e00-\u9fff]{2,}", source))

    for segment in re.findall(r"[\u4e00-\u9fff]{3,}", cleaned.replace(" ", "")):
        if len(segment) <= 6:
            terms.append(segment)
        for length in (5, 4, 3, 2):
            if len(segment) < length:
                continue
            for start in range(0, len(segment) - length + 1):
                terms.append(segment[start : start + length])

    filtered: list[str] = []
    for term in terms:
        term = normalize_text(term)
        if len(term) < 2 or term in RAG_QUERY_NOISE_WORDS:
            continue
        if term.isdigit():
            continue
        filtered.append(term)
    return expand_rag_terms(unique_keep_order(filtered))[:120]


def term_occurrences(text: str, term: str, match_case: bool = False) -> int:
    if not term:
        return 0
    haystack = text if match_case else text.casefold()
    needle = term if match_case else term.casefold()
    return haystack.count(needle)


def first_term_index(text: str, terms: Iterable[str], match_case: bool = False) -> tuple[int, str]:
    haystack = text if match_case else text.casefold()
    best_index = -1
    best_term = ""
    for term in terms:
        needle = term if match_case else term.casefold()
        index = haystack.find(needle)
        if index >= 0 and (best_index < 0 or index < best_index):
            best_index = index
            best_term = term
    return best_index, best_term


def focused_context_text(text: str, terms: Iterable[str], max_chars: int, match_case: bool = False) -> str:
    text = normalize_text(text)
    if max_chars <= 0:
        return ""
    if len(text) <= max_chars:
        return text

    local_index, _ = first_term_index(text, terms, match_case=match_case)
    if local_index < 0:
        return text[:max_chars]

    lines = text.splitlines()
    line_no = text[:local_index].count("\n")
    if lines and 0 <= line_no < len(lines):
        start_line = max(0, line_no - 2)
        end_line = min(len(lines), line_no + 3)
        while start_line < line_no or end_line > line_no + 1:
            snippet = "\n".join(lines[start_line:end_line]).strip()
            prefix = "...\n" if start_line > 0 else ""
            suffix = "\n..." if end_line < len(lines) else ""
            if len(prefix) + len(snippet) + len(suffix) <= max_chars:
                return prefix + snippet + suffix
            if end_line - line_no > line_no - start_line + 1 and end_line > line_no + 1:
                end_line -= 1
            elif start_line < line_no:
                start_line += 1
            else:
                break

    start = max(0, local_index - max_chars // 2)
    end = min(len(text), start + max_chars)
    if end - start < max_chars and start > 0:
        start = max(0, end - max_chars)

    prefix = "..." if start > 0 else ""
    suffix = "..." if end < len(text) else ""
    available = max_chars - len(prefix) - len(suffix)
    if available <= 0:
        return text[start:end][:max_chars]
    return prefix + text[start : start + available] + suffix


def score_rag_chunk(question: str, path: str, chunk: str, terms: list[str], match_case: bool = False) -> tuple[float, list[str]]:
    if not terms:
        return 0.0, []
    matched: list[str] = []
    score = 0.0
    cleaned_question = compact_question_text(question)
    chunk_cmp = chunk if match_case else chunk.casefold()
    path_cmp = path if match_case else path.casefold()

    phrase = cleaned_question if match_case else cleaned_question.casefold()
    if len(phrase) >= 3 and phrase in chunk_cmp:
        score += 80 + min(len(phrase), 30)

    for term in terms:
        term_cmp = term if match_case else term.casefold()
        chunk_count = chunk_cmp.count(term_cmp)
        path_count = path_cmp.count(term_cmp)
        if not chunk_count and not path_count:
            continue
        matched.append(term)
        length_weight = min(len(term), 8)
        score += min(chunk_count, 2) * (length_weight * length_weight)
        score += path_count * length_weight * 1.5
        if len(term) >= 4:
            score += 4

    unique_matched = unique_keep_order(matched)
    if unique_matched:
        positions = [chunk_cmp.find((term if match_case else term.casefold())) for term in unique_matched]
        positions = [pos for pos in positions if pos >= 0]
        if len(positions) >= 2:
            span = max(positions) - min(positions)
            if span <= 80:
                score += max(10.0, 55.0 - span * 0.7)
        coverage = len(unique_matched) / max(1, min(len(terms), 8))
        score *= 1.0 + min(coverage, 1.0) * 0.45
        if len(unique_matched) >= 2:
            score += 12
        if len(unique_matched) >= 3:
            score += 20
    return score, unique_matched


def score_filename_candidate(
    question: str,
    path: str,
    terms: list[str],
    match_case: bool = False,
) -> tuple[float, list[str]]:
    if not terms:
        return 0.0, []

    path_obj = Path(path)
    name = normalize_text(path_obj.name)
    stem = normalize_text(path_obj.stem)
    parent_text = normalize_text(" ".join(path_obj.parts[-4:-1]))
    full_text = normalize_text(f"{parent_text} {name}")
    separator_pattern = r"[\s_\-—–·.。/\\（）()【】\[\]《》<>]+"
    token_text = re.sub(separator_pattern, " ", full_text).strip()
    tokens = set(re.findall(r"[A-Za-z0-9_]{2,}|[\u4e00-\u9fff]{2,}", token_text))

    if not match_case:
        name_cmp = name.casefold()
        stem_cmp = stem.casefold()
        parent_cmp = parent_text.casefold()
        token_cmp = token_text.casefold()
        tokens_cmp = {token.casefold() for token in tokens}
    else:
        name_cmp = name
        stem_cmp = stem
        parent_cmp = parent_text
        token_cmp = token_text
        tokens_cmp = tokens

    joined_name = re.sub(r"\s+", "", name_cmp)
    joined_stem = re.sub(r"\s+", "", stem_cmp)
    compact_phrase = re.sub(r"\s+", "", compact_question_text(question))
    phrase_cmp = compact_phrase if match_case else compact_phrase.casefold()

    score = 0.0
    matched: list[str] = []
    if len(phrase_cmp) >= 3:
        if phrase_cmp in joined_stem:
            score += 120 + min(len(phrase_cmp), 20)
        elif phrase_cmp in joined_name:
            score += 95 + min(len(phrase_cmp), 20)

    for term in terms[:60]:
        term = normalize_text(term)
        if len(term) < 2 or term in RAG_QUERY_NOISE_WORDS:
            continue
        term_cmp = term if match_case else term.casefold()
        term_joined = re.sub(r"\s+", "", term_cmp)
        term_score = 0.0
        length_weight = min(len(term_joined), 8)
        if term_cmp in tokens_cmp:
            term_score += 34 + length_weight * 3
        if term_cmp in stem_cmp or term_joined in joined_stem:
            term_score += 30 + length_weight * 4
        if term_cmp in name_cmp or term_joined in joined_name:
            term_score += 24 + length_weight * 3
        if term_cmp in parent_cmp:
            term_score += 12 + length_weight * 1.5
        if term_cmp in token_cmp:
            term_score += 8 + length_weight
        if term_score > 0:
            matched.append(term)
            score += term_score

    unique_matched = unique_keep_order(matched)
    if unique_matched:
        coverage = len(unique_matched) / max(1, min(len(terms), 8))
        score *= 1.0 + min(coverage, 1.0) * 0.35
        if len(unique_matched) >= 2:
            score += 28
        if len(unique_matched) >= 3:
            score += 36
    ext = path_obj.suffix.lower()
    if ext in RAG_DOCUMENT_EXTENSIONS:
        score *= 1.12
    elif ext in RAG_LOW_PRIORITY_EXTENSIONS:
        score *= 0.45
    return score, unique_matched


def merge_file_candidates(
    limit: int,
    *groups: Iterable[FileNameCandidate],
) -> list[FileNameCandidate]:
    merged: dict[str, FileNameCandidate] = {}
    for group in groups:
        for item in group:
            existing = merged.get(item.path)
            if existing is None:
                merged[item.path] = item
                continue
            merged[item.path] = FileNameCandidate(
                path=item.path,
                score=existing.score + item.score,
                matched_terms=tuple(unique_keep_order([*existing.matched_terms, *item.matched_terms])[:8]),
            )
    ranked = list(merged.values())
    ranked.sort(key=lambda item: (-item.score, item.path))
    return ranked[:limit]


def rank_files_by_name(
    db_path: Path,
    question: str,
    match_case: bool = False,
    limit: int = 32,
) -> list[FileNameCandidate]:
    terms = rag_question_terms(question)
    if not terms or not db_path.exists() or limit <= 0:
        return []

    candidates: list[FileNameCandidate] = []
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute("SELECT path FROM files WHERE status = 'indexed' AND chunks > 0"):
            path = str(row["path"])
            score, matched_terms = score_filename_candidate(question, path, terms, match_case)
            if score <= 0:
                continue
            candidates.append(FileNameCandidate(path=path, score=score, matched_terms=tuple(matched_terms[:8])))
    finally:
        conn.close()

    candidates.sort(key=lambda item: (-item.score, item.path))
    return candidates[:limit]


def rank_files_by_content_preview(
    db_path: Path,
    question: str,
    match_case: bool = False,
    limit: int = 32,
) -> list[FileNameCandidate]:
    terms = rag_question_terms(question)
    if not terms or not db_path.exists() or limit <= 0:
        return []

    sql_terms = sorted(terms, key=len, reverse=True)[:24]
    like_parts = ["c.text LIKE ? ESCAPE '\\'" for _ in sql_terms]
    params = [f"%{escape_like(term)}%" for term in sql_terms]
    sql = f"""
        SELECT f.path, c.chunk_no, c.text
        FROM chunks c
        JOIN files f ON f.id = c.file_id
        WHERE f.status = 'indexed' AND ({" OR ".join(like_parts)})
        LIMIT 2400
    """

    scores: dict[str, float] = {}
    matched_by_path: dict[str, list[str]] = {}
    seen_chunks: set[tuple[str, int]] = set()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute(sql, params):
            path = str(row["path"])
            chunk_no = int(row["chunk_no"])
            key = (path, chunk_no)
            if key in seen_chunks:
                continue
            seen_chunks.add(key)
            chunk = normalize_text(str(row["text"]))
            chunk_score, matched_terms = score_rag_chunk(question, path, chunk, terms, match_case)
            if chunk_score <= 0:
                continue
            ext = Path(path).suffix.lower()
            if ext in RAG_DOCUMENT_EXTENSIONS:
                chunk_score *= 1.18
            elif ext in RAG_LOW_PRIORITY_EXTENSIONS:
                chunk_score *= 0.38
            scores[path] = scores.get(path, 0.0) + min(chunk_score, 900.0)
            if matched_terms:
                matched_by_path[path] = unique_keep_order([*matched_by_path.get(path, []), *matched_terms])[:8]
    finally:
        conn.close()

    candidates = [
        FileNameCandidate(path=path, score=score, matched_terms=tuple(matched_by_path.get(path, [])[:8]))
        for path, score in scores.items()
        if score > 0
    ]
    candidates.sort(key=lambda item: (-item.score, item.path))
    return candidates[:limit]


def rank_candidate_files(
    db_path: Path,
    question: str,
    match_case: bool = False,
    limit: int = 36,
) -> list[FileNameCandidate]:
    by_name = rank_files_by_name(db_path, question, match_case=match_case, limit=limit)
    by_content = rank_files_by_content_preview(db_path, question, match_case=match_case, limit=limit)
    return merge_file_candidates(limit, by_name, by_content)


def ranked_rag_contexts(
    db_path: Path,
    question: str,
    match_case: bool = False,
    limit: int = DEFAULT_RAG_TOP_K,
    file_paths: Optional[set[str]] = None,
    origin: str = "自动相关检索",
) -> list[RagContext]:
    terms = rag_question_terms(question)
    if not terms or not db_path.exists():
        return []
    filtered_paths = list(dict.fromkeys(file_paths or []))
    if file_paths is not None and not filtered_paths:
        return []

    sql_terms = sorted(terms, key=len, reverse=True)[:30]
    like_parts = ["c.text LIKE ? ESCAPE '\\'" for _ in sql_terms]
    like_parts.extend(["f.path LIKE ? ESCAPE '\\'" for _ in sql_terms[:12]])
    params = [f"%{escape_like(term)}%" for term in sql_terms]
    params.extend(f"%{escape_like(term)}%" for term in sql_terms[:12])
    path_filter = ""
    if filtered_paths:
        path_filter = f" AND f.path IN ({','.join(['?'] * len(filtered_paths))})"
        params.extend(filtered_paths)
    sql = f"""
        SELECT f.path, c.chunk_no, c.text
        FROM chunks c
        JOIN files f ON f.id = c.file_id
        WHERE ({" OR ".join(like_parts)}){path_filter}
        LIMIT 1800
    """

    scored: list[RagContext] = []
    seen: set[tuple[str, int]] = set()
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        for row in conn.execute(sql, params):
            path = str(row["path"])
            chunk_no = int(row["chunk_no"])
            key = (path, chunk_no)
            if key in seen:
                continue
            seen.add(key)
            chunk = normalize_text(str(row["text"]))
            score, matched_terms = score_rag_chunk(question, path, chunk, terms, match_case)
            if score <= 0:
                continue
            local_index, matched_keyword = first_term_index(chunk, matched_terms, match_case)
            source_location = infer_source_location(path, chunk, max(0, local_index), 1)
            if local_index < 0:
                source_location = SourceLocation("相关文件")
                matched_keyword = matched_terms[0] if matched_terms else ""
            scored.append(
                RagContext(
                    path=path,
                    keyword=matched_keyword,
                    position=source_location.label,
                    chunk_no=chunk_no,
                    text=chunk,
                    score=score,
                    matched_terms=tuple(matched_terms[:8]),
                    origin=origin,
                )
            )
    finally:
        conn.close()

    scored.sort(key=lambda item: (-item.score, item.path, item.chunk_no))
    return scored[:limit]


FOLLOWUP_MARKERS = {
    "它",
    "这个",
    "这些",
    "上述",
    "上面",
    "前面",
    "刚才",
    "继续",
    "还有",
    "那",
    "呢",
    "该",
    "其",
}


def is_followup_question(question: str) -> bool:
    compact = re.sub(r"\s+", "", normalize_text(question))
    if len(compact) <= 12:
        return True
    return any(marker in compact for marker in FOLLOWUP_MARKERS)


def recent_conversation_text(chat_history: Optional[list[dict[str, str]]], max_messages: int = 6) -> str:
    if not chat_history:
        return ""
    lines: list[str] = []
    for item in chat_history[-max_messages:]:
        role = item.get("role", "")
        content = normalize_text(item.get("content", ""))
        if not content:
            continue
        label = "用户" if role == "user" else "助手"
        lines.append(f"{label}：{content[:1200]}")
    return "\n".join(lines)


def retrieval_query_for_question(question: str, chat_history: Optional[list[dict[str, str]]] = None) -> str:
    question = normalize_text(question)
    if not chat_history or not is_followup_question(question):
        return question
    previous_questions = [
        normalize_text(item.get("content", ""))
        for item in chat_history
        if item.get("role") == "user" and normalize_text(item.get("content", ""))
    ]
    recent_questions = previous_questions[-2:]
    return normalize_text(" ".join([*recent_questions, question]))


def should_skip_local_retrieval(question: str, chat_history: Optional[list[dict[str, str]]] = None) -> bool:
    if not chat_history:
        return False
    text = re.sub(r"\s+", "", normalize_text(question))
    if not text:
        return False
    retrieval_intent = (
        "重新检索",
        "重新搜索",
        "重新查",
        "现在查",
        "帮我查",
        "请查",
        "查找",
        "搜索资料",
        "检索资料",
        "在资料",
        "在文件",
        "根据资料",
        "引用资料",
    )
    if any(intent in text for intent in retrieval_intent):
        return False
    context_patterns = (
        r"(刚才|上次|前面|上一轮|这次|此次|第一次|第二次).*(回答|对话|检索|搜索|找到|查到|引用|资料|结果)",
        r"(为什么|为何).*(没|没有|未).*(检索|搜索|找到|查到)",
        r"(为什么|为何).*(检索|搜索|引用).*(不|没|没有|错误|错|偏)",
        r"(我|用户).*(指定|明确).*(文件|资料|文档)",
        r"(你|系统|程序).*(需要|应该|可以).*(检索|搜索)",
        r"(不用|不需要|无需).*(检索|搜索)",
        r"(直接|只).*(根据|结合).*(上下文|对话|聊天记录)",
        r"(这|这种|上述).*(情况|问题|现象).*(为什么|原因|说明)",
        r"(总结|解释|分析).*(刚才|前面|上一轮|这次|两次).*(对话|问答|回答|检索)",
    )
    if any(re.search(pattern, text) for pattern in context_patterns):
        return True
    return False


def build_context_only_prompt(question: str, chat_history: Optional[list[dict[str, str]]] = None) -> str:
    conversation = recent_conversation_text(chat_history, max_messages=8)
    parts = [
        "任务：根据最近对话回答用户当前问题。",
        "当前问题看起来是在询问对话过程、检索策略、上一轮回答或软件行为，不需要调用本地资料库。",
        "请只依据最近对话和用户当前问题进行解释；不要编造本地文件内容；如果需要重新检索才能确认，就明确说明。",
        "回答要直接、简洁，并给出可执行的改进建议。",
    ]
    if conversation:
        parts.extend(["", "最近对话：", conversation])
    parts.extend(["", f"当前问题：{question}"])
    return "\n".join(parts)


def escape_like(term: str) -> str:
    return term.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


def find_all(haystack: str, needle: str, match_case: bool) -> Iterator[int]:
    if not needle:
        return
    if match_case:
        search_haystack = haystack
        search_needle = needle
    else:
        search_haystack = haystack.casefold()
        search_needle = needle.casefold()
    start = 0
    while True:
        index = search_haystack.find(search_needle, start)
        if index == -1:
            break
        yield index
        start = index + max(1, len(search_needle))


def chunk_contains(chunk: str, term: str, match_case: bool) -> bool:
    if match_case:
        return term in chunk
    return term.casefold() in chunk.casefold()


def location_in_chunk(start_line: int, start_col: int, chunk: str, local_index: int) -> tuple[int, int]:
    prefix = chunk[:local_index]
    line_delta = prefix.count("\n")
    if line_delta == 0:
        return start_line, start_col + local_index
    last_newline = prefix.rfind("\n")
    return start_line + line_delta, local_index - last_newline


def make_snippet(text: str, index: int, keyword: str, radius: int = 70) -> str:
    start = max(0, index - radius)
    end = min(len(text), index + len(keyword) + radius)
    snippet = text[start:end].replace("\n", " ")
    snippet = re.sub(r"\s+", " ", snippet).strip()
    return snippet


def mark_keyword_in_text(text: str, keyword: str, match_case: bool = False) -> str:
    if not keyword:
        return text
    flags = 0 if match_case else re.IGNORECASE
    return re.sub(re.escape(keyword), lambda m: f"【{m.group(0)}】", text, flags=flags)


def line_containing_index(text: str, index: int) -> str:
    start = text.rfind("\n", 0, index) + 1
    end = text.find("\n", index)
    if end == -1:
        end = len(text)
    return text[start:end].strip()


def infer_page_from_chunk(chunk: str, local_index: int) -> int:
    before = chunk[: local_index + 1]
    matches = list(re.finditer(r"\[Page\s+(\d+)\]", before, flags=re.IGNORECASE))
    if matches:
        return int(matches[-1].group(1))
    after = chunk[local_index:]
    match = re.search(r"\[Page\s+(\d+)\]", after, flags=re.IGNORECASE)
    return int(match.group(1)) if match else 0


def infer_excel_location_from_line(line: str) -> tuple[str, str]:
    match = re.search(r"\[Sheet\s+\d+:\s*([^\]]+)\]\s+([A-Z]{1,4}\d+):", line)
    if match:
        return match.group(1).strip(), match.group(2).strip()
    match = re.search(r"\[Sheet\s+(\d+)\]\s+([A-Z]{1,4}\d+):", line)
    if match:
        return f"Sheet {match.group(1)}", match.group(2).strip()
    return "", ""


def infer_doc_heading(chunk: str, local_index: int) -> str:
    before = chunk[:local_index]
    candidates = [line.strip() for line in before.splitlines() if line.strip()]
    for line in reversed(candidates[-12:]):
        if len(line) <= 80 and not line.endswith(("。", ".", "，", ",", ";", "；")):
            return line
    return ""


def infer_source_location(path: str, chunk: str, local_index: int, occurrence: int) -> SourceLocation:
    ext = Path(path).suffix.lower()
    if ext in {".xlsx", ".xlsm", ".xls"}:
        sheet, cell = infer_excel_location_from_line(line_containing_index(chunk, local_index))
        if sheet and cell:
            return SourceLocation(f"工作表：{sheet}，单元格：{cell}", sheet=sheet, cell=cell)
        return SourceLocation(f"Excel 第 {occurrence} 次命中")

    if ext == ".pdf":
        page = infer_page_from_chunk(chunk, local_index)
        if page:
            return SourceLocation(f"PDF 第 {page} 页", page=page)
        return SourceLocation(f"PDF 第 {occurrence} 次命中")

    if ext == ".docx":
        heading = infer_doc_heading(chunk, local_index)
        if heading:
            return SourceLocation(f"Word 章节附近：{heading}")
        return SourceLocation(f"Word 第 {occurrence} 次命中")

    if ext == ".pptx":
        slide = infer_page_from_chunk(chunk.replace("[Slide", "[Page"), local_index)
        if slide:
            return SourceLocation(f"PPT 第 {slide} 页")
        return SourceLocation(f"第 {occurrence} 次命中")

    return SourceLocation(f"第 {occurrence} 次命中")


def search_index(
    db_path: Path,
    query: str,
    require_all_terms: bool = False,
    match_case: bool = False,
    phrase: bool = False,
    limit: int = 500,
    file_paths: Optional[set[str]] = None,
) -> list[SearchHit]:
    terms = parse_keywords(query, phrase=phrase)
    if not terms:
        return []
    if not db_path.exists():
        raise ValueError(f"Index database does not exist: {db_path}")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    filtered_paths = list(dict.fromkeys(file_paths or []))
    if file_paths is not None and not filtered_paths:
        conn.close()
        return []
    where = " OR ".join(["c.text LIKE ? ESCAPE '\\'"] * len(terms))
    params = [f"%{escape_like(term)}%" for term in terms]
    path_filter = ""
    if filtered_paths:
        path_filter = f" AND f.path IN ({','.join(['?'] * len(filtered_paths))})"
        params.extend(filtered_paths)
    sql = f"""
        SELECT f.path, c.chunk_no, c.start_offset, c.start_line, c.start_col, c.text
        FROM chunks c
        JOIN files f ON f.id = c.file_id
        WHERE ({where}){path_filter}
        ORDER BY f.path, c.chunk_no
    """

    hits: list[SearchHit] = []
    seen: set[tuple[str, str, int]] = set()
    occurrence_counts: dict[tuple[str, str], int] = {}
    try:
        for row in conn.execute(sql, params):
            chunk = str(row["text"])
            if require_all_terms and not all(chunk_contains(chunk, term, match_case) for term in terms):
                continue

            for term in terms:
                if require_all_terms is False and not chunk_contains(chunk, term, match_case):
                    continue
                for local_index in find_all(chunk, term, match_case):
                    global_offset = int(row["start_offset"]) + local_index
                    key = (str(row["path"]), term.casefold() if not match_case else term, global_offset)
                    if key in seen:
                        continue
                    seen.add(key)
                    occurrence_key = (str(row["path"]), term.casefold() if not match_case else term)
                    occurrence = occurrence_counts.get(occurrence_key, 0) + 1
                    occurrence_counts[occurrence_key] = occurrence
                    line, col = location_in_chunk(int(row["start_line"]), int(row["start_col"]), chunk, local_index)
                    source_location = infer_source_location(str(row["path"]), chunk, local_index, occurrence)
                    hits.append(
                        SearchHit(
                            path=str(row["path"]),
                            keyword=term,
                            position=source_location.label,
                            line=line,
                            column=col,
                            offset=global_offset,
                            chunk_no=int(row["chunk_no"]),
                            snippet=make_snippet(chunk, local_index, term),
                            occurrence=occurrence,
                            page=source_location.page,
                            sheet=source_location.sheet,
                            cell=source_location.cell,
                        )
                    )
                    if len(hits) >= limit:
                        return hits
    finally:
        conn.close()
    return hits


def hits_to_rag_contexts(
    db_path: Path,
    hits: list[SearchHit],
    limit: int = DEFAULT_RAG_TOP_K,
    origin: str = "检索结果",
) -> list[RagContext]:
    if not hits or not db_path.exists() or limit <= 0:
        return []

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    contexts: list[RagContext] = []
    seen: set[tuple[str, int]] = set()
    try:
        for hit in hits:
            key = (hit.path, hit.chunk_no)
            if key in seen:
                continue
            seen.add(key)
            row = conn.execute(
                """
                SELECT c.text
                FROM chunks c
                JOIN files f ON f.id = c.file_id
                WHERE f.path = ? AND c.chunk_no = ?
                """,
                (hit.path, hit.chunk_no),
            ).fetchone()
            chunk_text = normalize_text(str(row["text"])) if row else hit.snippet
            contexts.append(
                RagContext(
                    path=hit.path,
                    keyword=hit.keyword,
                    position=hit.position,
                    chunk_no=hit.chunk_no,
                    text=chunk_text,
                    score=10000.0 - len(contexts),
                    matched_terms=(hit.keyword,),
                    origin=origin,
                )
            )
            if len(contexts) >= limit:
                break
    finally:
        conn.close()
    return contexts


def merge_rag_contexts(limit: int, *groups: list[RagContext]) -> list[RagContext]:
    merged: list[RagContext] = []
    seen: set[tuple[str, int]] = set()
    high_priority_paths: set[str] = set()
    for group in groups:
        for context in group:
            key = (context.path, context.chunk_no)
            if key in seen:
                continue
            if context.origin == "自动相关检索" and context.path in high_priority_paths:
                continue
            seen.add(key)
            merged.append(context)
            if context.origin in {"当前问题精确检索", "当前结果栏", "精确关键词匹配", "文件名候选精确检索", "文件名候选相关检索"}:
                high_priority_paths.add(context.path)
            if len(merged) >= limit:
                return merged
    return merged


def exact_rag_contexts(
    db_path: Path,
    query: str,
    require_all_terms: bool = False,
    match_case: bool = False,
    phrase: bool = False,
    limit: int = DEFAULT_RAG_TOP_K,
    file_paths: Optional[set[str]] = None,
    origin: str = "精确关键词匹配",
) -> list[RagContext]:
    """Recall exact keyword hits, then rank them instead of returning the first path order.

    The old logic stopped as soon as any candidate term produced hits. In a large
    knowledge base this often returns alphabetically earlier but weakly-related
    files. This version gathers hits from multiple query candidates and scores
    the resulting chunks against the whole question.
    """
    if limit <= 0:
        return []
    hits: list[SearchHit] = []
    seen_hits: set[tuple[str, str, int]] = set()
    max_hits = max(limit * 20, 80)
    per_candidate_limit = max(limit * 6, 24)
    for candidate in rag_search_candidates(query)[:36]:
        candidate_hits = search_index(
            db_path,
            candidate,
            require_all_terms=require_all_terms,
            match_case=match_case,
            phrase=phrase,
            limit=per_candidate_limit,
            file_paths=file_paths,
        )
        for hit in candidate_hits:
            key = (hit.path, hit.keyword.casefold() if not match_case else hit.keyword, hit.offset)
            if key in seen_hits:
                continue
            seen_hits.add(key)
            hits.append(hit)
            if len(hits) >= max_hits:
                break
        if len(hits) >= max_hits:
            break

    contexts = hits_to_rag_contexts(db_path, hits, limit=max_hits, origin=origin)
    terms = rag_question_terms(query)
    for context in contexts:
        score, matched_terms = score_rag_chunk(query, context.path, context.text, terms, match_case)
        exact_bonus = 120.0 if context.keyword else 0.0
        context.score = score + exact_bonus
        if matched_terms:
            context.matched_terms = tuple(unique_keep_order([context.keyword, *matched_terms])[:8])
    contexts.sort(key=lambda item: (-item.score, item.path, item.chunk_no))
    return contexts[:limit]


def normalize_chat_api_url(api_url: str) -> str:
    api_url = api_url.strip().rstrip("/")
    if not api_url:
        raise ValueError("请填写 API 地址。")
    if api_url.endswith("/chat/completions"):
        return api_url
    if api_url.endswith("/v1"):
        return api_url + "/chat/completions"
    if "api.openai.com" in api_url:
        return api_url + "/v1/chat/completions"
    return api_url + "/chat/completions"


def load_rag_settings() -> dict[str, object]:
    settings = {
        "api_url": os.environ.get("ENTERPRISE_RETRIEVAL_API_URL", DEFAULT_RAG_API_URL),
        "model": os.environ.get("ENTERPRISE_RETRIEVAL_MODEL", DEFAULT_RAG_MODEL),
        "api_key": os.environ.get("ENTERPRISE_RETRIEVAL_API_KEY", os.environ.get("OPENAI_API_KEY", "")),
        "top_k": DEFAULT_RAG_TOP_K,
        "save_key": False,
    }
    path = default_settings_path()
    if path.exists():
        try:
            stored = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(stored, dict):
                settings.update(stored)
        except Exception:
            pass
    return settings


def save_rag_settings(api_url: str, model: str, api_key: str, top_k: int, save_key: bool) -> None:
    path = default_settings_path()
    path.parent.mkdir(parents=True, exist_ok=True)
    data: dict[str, object] = {
        "api_url": api_url.strip(),
        "model": model.strip(),
        "top_k": int(top_k),
        "save_key": bool(save_key),
    }
    if save_key:
        data["api_key"] = api_key
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def candidate_file_contexts(
    db_path: Path,
    question: str,
    file_candidates: list[FileNameCandidate],
    match_case: bool = False,
    limit: int = DEFAULT_RAG_TOP_K,
    chunks_per_file: int = 2,
    max_candidate_files: int = 12,
    origin: str = "文件名/文件夹候选",
) -> list[RagContext]:
    """Return representative chunks from files whose path/name is highly relevant.

    This is a fallback channel for cases where the right file/folder can be
    inferred from the question, but the exact wording inside the document does
    not match the user's wording.
    """
    if not file_candidates or not db_path.exists() or limit <= 0:
        return []

    terms = rag_question_terms(question)
    scored: list[RagContext] = []
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        for file_rank, candidate in enumerate(file_candidates[:max_candidate_files]):
            rows = conn.execute(
                """
                SELECT f.path, c.chunk_no, c.text
                FROM chunks c
                JOIN files f ON f.id = c.file_id
                WHERE f.path = ? AND f.status = 'indexed'
                ORDER BY c.chunk_no
                LIMIT 30
                """,
                (candidate.path,),
            ).fetchall()
            local_items: list[RagContext] = []
            for row in rows:
                path = str(row["path"])
                chunk_no = int(row["chunk_no"])
                chunk = normalize_text(str(row["text"]))
                chunk_score, matched_terms = score_rag_chunk(question, path, chunk, terms, match_case)
                # Keep early chunks as a weak fallback even when wording does not match.
                early_bonus = max(0.0, 18.0 - chunk_no * 2.0)
                candidate_bonus = max(0.0, candidate.score * 0.22 - file_rank * 3.0)
                final_score = chunk_score + early_bonus + candidate_bonus
                local_index, matched_keyword = first_term_index(chunk, matched_terms or candidate.matched_terms, match_case)
                if local_index >= 0:
                    source_location = infer_source_location(path, chunk, local_index, 1)
                else:
                    source_location = SourceLocation("文件名/文件夹相关，内容候选")
                    matched_keyword = next(iter(candidate.matched_terms), "")
                local_items.append(
                    RagContext(
                        path=path,
                        keyword=matched_keyword,
                        position=source_location.label,
                        chunk_no=chunk_no,
                        text=chunk,
                        score=final_score,
                        matched_terms=tuple(unique_keep_order([*candidate.matched_terms, *matched_terms])[:8]),
                        origin=origin,
                    )
                )
            local_items.sort(key=lambda item: (-item.score, item.chunk_no))
            scored.extend(local_items[:chunks_per_file])
    finally:
        conn.close()

    scored.sort(key=lambda item: (-item.score, item.path, item.chunk_no))
    return scored[:limit]


def fetch_rag_contexts(
    db_path: Path,
    query: str,
    require_all_terms: bool = False,
    match_case: bool = False,
    phrase: bool = False,
    limit: int = DEFAULT_RAG_TOP_K,
    chat_history: Optional[list[dict[str, str]]] = None,
    priority_hits: Optional[list[SearchHit]] = None,
    preferred_file_paths: Optional[set[str]] = None,
) -> list[RagContext]:
    retrieval_query = retrieval_query_for_question(query, chat_history)
    over_limit = max(limit * 3, limit)

    priority_contexts = hits_to_rag_contexts(
        db_path,
        priority_hits or [],
        limit=over_limit,
        origin="当前问题精确检索",
    )
    terms = rag_question_terms(retrieval_query)
    for context in priority_contexts:
        score, matched_terms = score_rag_chunk(retrieval_query, context.path, context.text, terms, match_case)
        context.score = score + 150.0
        if matched_terms:
            context.matched_terms = tuple(unique_keep_order([context.keyword, *matched_terms])[:8])
    priority_contexts.sort(key=lambda item: (-item.score, item.path, item.chunk_no))
    priority_contexts = priority_contexts[:limit]

    file_candidates = rank_candidate_files(
        db_path,
        retrieval_query,
        match_case=match_case,
        limit=max(limit * 8, 32),
    )
    preferred_paths = set(preferred_file_paths or set())
    if preferred_paths:
        existing = {item.path for item in file_candidates}
        for path in preferred_paths:
            if path not in existing:
                file_candidates.append(FileNameCandidate(path=path, score=80.0, matched_terms=()))
        file_candidates.sort(key=lambda item: (-item.score, item.path))
    candidate_paths = {item.path for item in file_candidates}

    filename_exact_contexts: list[RagContext] = []
    filename_ranked_contexts: list[RagContext] = []
    filename_broad_contexts: list[RagContext] = []
    if candidate_paths:
        filename_exact_contexts = exact_rag_contexts(
            db_path,
            retrieval_query,
            require_all_terms=require_all_terms,
            match_case=match_case,
            phrase=phrase,
            limit=over_limit,
            file_paths=candidate_paths,
            origin="文件名候选精确检索",
        )
        filename_ranked_contexts = ranked_rag_contexts(
            db_path,
            retrieval_query,
            match_case=match_case,
            limit=over_limit,
            file_paths=candidate_paths,
            origin="文件名候选相关检索",
        )
        filename_broad_contexts = candidate_file_contexts(
            db_path,
            retrieval_query,
            file_candidates,
            match_case=match_case,
            limit=over_limit,
            origin="文件名/文件夹候选",
        )

    exact_contexts = exact_rag_contexts(
        db_path,
        retrieval_query,
        require_all_terms=require_all_terms,
        match_case=match_case,
        phrase=phrase,
        limit=over_limit,
        origin="全库精确关键词匹配",
    )
    ranked_contexts = ranked_rag_contexts(
        db_path,
        retrieval_query,
        match_case=match_case,
        limit=over_limit,
        origin="全库相关检索",
    )

    # Merge by channel priority, but over-recall first so a weak early channel does not
    # completely hide stronger full-library results.
    merged = merge_rag_contexts(
        over_limit,
        priority_contexts,
        filename_exact_contexts,
        exact_contexts,
        filename_ranked_contexts,
        ranked_contexts,
        filename_broad_contexts,
    )
    merged.sort(
        key=lambda item: (
            -(
                item.score
                + (80 if item.origin in {"当前问题精确检索", "文件名候选精确检索", "全库精确关键词匹配"} else 0)
                + (35 if item.origin == "文件名/文件夹候选" else 0)
            ),
            item.path,
            item.chunk_no,
        )
    )
    return merged[:limit]


def build_rag_prompt(
    question: str,
    contexts: list[RagContext],
    max_context_chars: int = DEFAULT_RAG_MAX_CONTEXT_CHARS,
    chat_history: Optional[list[dict[str, str]]] = None,
) -> str:
    conversation = recent_conversation_text(chat_history)
    parts = [
        "任务：回答用户的问题，不是总结关键词。",
        "先给直接结论，再给依据；回答要简洁、自然，不要泛泛介绍无关背景。",
        "请先判断检索片段是否能直接支持答案；如果证据不足，要说明已经找到哪些线索、还缺少哪类直接依据。",
        "如果当前问题是追问，请结合“最近对话”理解代词和省略内容，但依据仍必须来自本地检索结果。",
        "证据来源为“当前问题精确检索”“文件名候选精确检索”和“精确关键词匹配”的片段优先级最高，要优先据此回答。",
        "如果片段或位置中出现用户问题里的人名、编号、制度名等精确词，不要直接说没有找到；应说明资料里出现了什么，以及是否足以判断身份或结论。",
        "每个关键事实或判断后面都要使用 [1]、[2] 这样的编号引用本地来源；没有来源支撑的内容要明确标为推断或不要写。",
    ]
    if conversation:
        parts.extend(["", "最近对话：", conversation])
    parts.extend(["", f"当前问题：{question}", "", "本地检索结果："])
    used_chars = 0
    question_terms = rag_question_terms(question)
    for index, context in enumerate(contexts, start=1):
        matched = "、".join(context.matched_terms) if context.matched_terms else context.keyword
        origin = context.origin or "本地检索"
        header = f"[{index}] 证据来源：{origin}\n文件：{context.path}\n位置：{context.position}\n匹配要点：{matched}\n内容："
        remaining = max_context_chars - used_chars - len(header)
        if remaining <= 0:
            break
        focus_terms = unique_keep_order([context.keyword, *context.matched_terms, *question_terms])
        text = focused_context_text(context.text, focus_terms, remaining)
        used_chars += len(header) + len(text)
        parts.append(header + text)
    return "\n\n".join(parts)


def call_chat_completion(
    api_url: str,
    api_key: str,
    model: str,
    question: str,
    contexts: list[RagContext],
    chat_history: Optional[list[dict[str, str]]] = None,
    timeout: int = 90,
) -> str:
    if not model.strip():
        raise ValueError("请填写模型名称。")
    if not contexts:
        raise ValueError("没有找到可用于问答的本地检索结果。请先确认索引已建立，或把问题里的核心词写得更具体，例如直接输入“轮毂”。")

    url = normalize_chat_api_url(api_url)
    prompt = build_rag_prompt(question, contexts, chat_history=chat_history)
    payload = {
        "model": model.strip(),
        "messages": [
            {
                "role": "system",
                "content": "你是企业本地文档问答助手。只根据提供的本地检索片段回答，优先使用当前问题精确检索和精确关键词匹配的证据。能直接回答就先给结论；证据不足时说明已找到的线索和缺少的直接依据。所有关键事实都要标注来源编号。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "stream": False,
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if api_key.strip():
        headers["Authorization"] = f"Bearer {api_key.strip()}"
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"API 请求失败：HTTP {exc.code} {body[:800]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"API 连接失败：{exc}") from exc

    data = json.loads(raw)
    choices = data.get("choices")
    if isinstance(choices, list) and choices:
        message = choices[0].get("message") if isinstance(choices[0], dict) else None
        if isinstance(message, dict):
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return content.strip()
        text = choices[0].get("text") if isinstance(choices[0], dict) else None
        if isinstance(text, str) and text.strip():
            return text.strip()
    output_text = data.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()
    raise RuntimeError("API 返回成功，但没有解析到回答内容。")


def call_context_completion(
    api_url: str,
    api_key: str,
    model: str,
    question: str,
    chat_history: Optional[list[dict[str, str]]] = None,
    timeout: int = 90,
) -> str:
    if not model.strip():
        raise ValueError("请填写模型名称。")

    url = normalize_chat_api_url(api_url)
    prompt = build_context_only_prompt(question, chat_history=chat_history)
    payload = {
        "model": model.strip(),
        "messages": [
            {
                "role": "system",
                "content": "你是企业知识库助手。当前问题不需要检索本地资料库，请根据最近对话和用户当前问题回答。不要编造本地文件内容；如果需要重新检索才能确认，要明确说明。",
            },
            {"role": "user", "content": prompt},
        ],
        "temperature": 0.2,
        "stream": False,
    }
    headers = {
        "Content-Type": "application/json",
        "Accept": "application/json",
    }
    if api_key.strip():
        headers["Authorization"] = f"Bearer {api_key.strip()}"
    request = urllib.request.Request(
        url,
        data=json.dumps(payload, ensure_ascii=False).encode("utf-8"),
        headers=headers,
        method="POST",
    )
    try:
        with urllib.request.urlopen(request, timeout=timeout) as response:
            raw = response.read().decode("utf-8", errors="replace")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="replace")
        raise RuntimeError(f"API 请求失败：HTTP {exc.code} {body[:800]}") from exc
    except urllib.error.URLError as exc:
        raise RuntimeError(f"API 连接失败：{exc}") from exc

    data = json.loads(raw)
    choices = data.get("choices")
    if isinstance(choices, list) and choices:
        message = choices[0].get("message") if isinstance(choices[0], dict) else None
        if isinstance(message, dict):
            content = message.get("content")
            if isinstance(content, str) and content.strip():
                return content.strip()
        text = choices[0].get("text") if isinstance(choices[0], dict) else None
        if isinstance(text, str) and text.strip():
            return text.strip()
    output_text = data.get("output_text")
    if isinstance(output_text, str) and output_text.strip():
        return output_text.strip()
    raise RuntimeError("API 返回成功，但没有解析到回答内容。")


def export_hits_csv(path: Path, hits: Iterable[SearchHit]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerow(["file_path", "keyword", "position", "snippet"])
        for hit in hits:
            writer.writerow([hit.path, hit.keyword, hit.position, hit.snippet])


def print_hits_table(hits: Iterable[SearchHit]) -> None:
    for hit in hits:
        print(f"{hit.path} [{hit.keyword} | {hit.position}] {hit.snippet}")


def run_powershell(script: str, *args: str) -> subprocess.Popen:
    creationflags = getattr(subprocess, "CREATE_NO_WINDOW", 0)
    wrapped_script = "& {\n" + script + "\n}"
    return subprocess.Popen(
        ["powershell.exe", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", wrapped_script, *args],
        creationflags=creationflags,
    )


def open_default(path: str) -> None:
    os.startfile(path)  # type: ignore[attr-defined]


def open_word_at_keyword(hit: SearchHit) -> None:
    script = r"""
param([string]$FilePath, [string]$Keyword, [string]$OccurrenceText)
$ErrorActionPreference = 'Stop'
try {
    $occurrence = [Math]::Max(1, [int]$OccurrenceText)
    $word = New-Object -ComObject Word.Application
    $word.Visible = $true
    $doc = $word.Documents.Open($FilePath)
    $selection = $word.Selection
    $null = $selection.HomeKey(6)
    for ($i = 1; $i -le $occurrence; $i++) {
        $selection.Find.ClearFormatting()
        $selection.Find.Text = $Keyword
        $selection.Find.Forward = $true
        $selection.Find.Wrap = 0
        $found = $selection.Find.Execute()
        if (-not $found) { break }
    }
    $word.Activate()
} catch {
    Start-Process -LiteralPath $FilePath
}
"""
    run_powershell(script, hit.path, hit.keyword, str(max(1, hit.occurrence)))


def open_excel_at_cell(hit: SearchHit) -> None:
    script = r"""
param([string]$FilePath, [string]$SheetName, [string]$CellAddress, [string]$Keyword)
$ErrorActionPreference = 'Stop'
try {
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $true
    $workbook = $excel.Workbooks.Open($FilePath)
    $target = $null
    if ($SheetName -and $CellAddress) {
        try {
            $worksheet = $workbook.Worksheets.Item($SheetName)
            $worksheet.Activate()
            $target = $worksheet.Range($CellAddress)
        } catch {
            $target = $null
        }
    }
    if ($null -eq $target -and $Keyword) {
        foreach ($worksheet in $workbook.Worksheets) {
            $range = $worksheet.UsedRange
            $found = $range.Find($Keyword)
            if ($null -ne $found) {
                $worksheet.Activate()
                $target = $found
                break
            }
        }
    }
    if ($null -ne $target) {
        $target.Select()
        $excel.Goto($target, $true)
    }
    $excel.Activate()
} catch {
    Start-Process -LiteralPath $FilePath
}
"""
    run_powershell(script, hit.path, hit.sheet, hit.cell, hit.keyword)


def open_pdf_at_page(hit: SearchHit) -> None:
    if hit.page > 0:
        try:
            uri = Path(hit.path).resolve().as_uri() + f"#page={hit.page}"
            os.startfile(uri)  # type: ignore[attr-defined]
            return
        except Exception:
            pass
    open_default(hit.path)


def open_hit_target(hit: SearchHit) -> None:
    ext = Path(hit.path).suffix.lower()
    if ext == ".pdf":
        open_pdf_at_page(hit)
    else:
        open_default(hit.path)


def launch_gui() -> None:
    import tkinter as tk
    import tkinter.font as tkfont
    from tkinter import filedialog, messagebox, ttk

    COLORS = {
        "app_bg": "#F3F5F8",
        "panel_bg": "#F7F9FC",
        "sidebar_bg": "#080E1A",
        "sidebar_card": "#0F172A",
        "sidebar_hover": "#152238",
        "sidebar_muted": "#8B9AB0",
        "sidebar_text": "#F8FAFC",
        "card": "#FFFFFF",
        "card_alt": "#F7F9FC",
        "line": "#E5EAF1",
        "line_strong": "#CBD5E1",
        "shadow": "#E8EDF5",
        "text": "#111827",
        "muted": "#667085",
        "muted_light": "#98A2B3",
        "primary": "#2563EB",
        "primary_dark": "#1D4ED8",
        "primary_soft": "#EDF4FF",
        "success": "#047857",
        "warning": "#B45309",
        "danger": "#B91C1C",
        "bubble_user": "#EFF6FF",
        "bubble_ai": "#FFFFFF",
    }
    RADIUS = {
        "card": 18,
        "button": 12,
        "input": 20,
        "bubble": 18,
        "citation": 14,
    }

    def short_path(path: str, max_len: int = 34) -> str:
        path = path.strip()
        if not path:
            return "未选择"
        if len(path) <= max_len:
            return path
        return "..." + path[-max_len:]

    def draw_rounded_rect(
        canvas: tk.Canvas,
        x1: int,
        y1: int,
        x2: int,
        y2: int,
        radius: int,
        *,
        fill: str,
        outline: str = "",
        width: int = 1,
    ) -> None:
        radius = max(1, min(radius, max(1, (x2 - x1) // 2), max(1, (y2 - y1) // 2)))
        canvas.create_rectangle(x1 + radius, y1, x2 - radius, y2, fill=fill, outline=fill)
        canvas.create_rectangle(x1, y1 + radius, x2, y2 - radius, fill=fill, outline=fill)
        canvas.create_oval(x1, y1, x1 + radius * 2, y1 + radius * 2, fill=fill, outline=fill)
        canvas.create_oval(x2 - radius * 2, y1, x2, y1 + radius * 2, fill=fill, outline=fill)
        canvas.create_oval(x1, y2 - radius * 2, x1 + radius * 2, y2, fill=fill, outline=fill)
        canvas.create_oval(x2 - radius * 2, y2 - radius * 2, x2, y2, fill=fill, outline=fill)
        if outline:
            canvas.create_arc(x1, y1, x1 + radius * 2, y1 + radius * 2, start=90, extent=90, style=tk.ARC, outline=outline, width=width)
            canvas.create_arc(x2 - radius * 2, y1, x2, y1 + radius * 2, start=0, extent=90, style=tk.ARC, outline=outline, width=width)
            canvas.create_arc(x2 - radius * 2, y2 - radius * 2, x2, y2, start=270, extent=90, style=tk.ARC, outline=outline, width=width)
            canvas.create_arc(x1, y2 - radius * 2, x1 + radius * 2, y2, start=180, extent=90, style=tk.ARC, outline=outline, width=width)
            canvas.create_line(x1 + radius, y1, x2 - radius, y1, fill=outline, width=width)
            canvas.create_line(x2, y1 + radius, x2, y2 - radius, fill=outline, width=width)
            canvas.create_line(x1 + radius, y2, x2 - radius, y2, fill=outline, width=width)
            canvas.create_line(x1, y1 + radius, x1, y2 - radius, fill=outline, width=width)

    class ModernButton(tk.Frame):
        def __init__(
            self,
            master: tk.Misc,
            text: str,
            command: Optional[Callable[[], None]] = None,
            *,
            bg: str = COLORS["card_alt"],
            fg: str = COLORS["text"],
            hover_bg: str = "#EEF4FF",
            active_bg: str = "#E0EAFF",
            border: str = "",
            hover_border: str = "",
            active_border: str = "",
            font: Optional[tkfont.Font] = None,
            padx: int = 16,
            pady: int = 9,
            anchor: str = "center",
            radius: int = RADIUS["button"],
            min_width: Optional[int] = None,
        ) -> None:
            super().__init__(master, bg=bg, bd=0, highlightthickness=0)
            self.text = text
            self.command = command
            self.normal_bg = bg
            self.hover_bg = hover_bg
            self.active_bg = active_bg
            self.border = border
            self.hover_border = hover_border or border
            self.active_border = active_border or self.hover_border
            self.fg = fg
            self.disabled_bg = "#E5E7EB" if bg == COLORS["primary"] else "#EEF2F7"
            self.disabled_fg = "#94A3B8"
            self.font = font
            self.padx = padx
            self.pady = pady
            self.anchor = anchor
            self.radius = radius
            self.state = tk.NORMAL
            self._pressed = False
            line_height = (font.metrics("linespace") if font else 16) + pady * 2
            text_width = (font.measure(text) if font else max(28, len(text) * 9)) + padx * 2 + 10
            button_width = max(min_width or 0, text_width, 38)
            self.canvas = tk.Canvas(
                self,
                width=button_width,
                height=max(38, line_height),
                bg=self._parent_bg(),
                bd=0,
                highlightthickness=0,
                cursor="hand2",
            )
            self.canvas.pack(fill=tk.BOTH, expand=True)
            self.canvas.bind("<Configure>", lambda _event: self._draw())
            self.canvas.bind("<Enter>", self._on_enter)
            self.canvas.bind("<Leave>", self._on_leave)
            self.canvas.bind("<ButtonPress-1>", self._on_press)
            self.canvas.bind("<ButtonRelease-1>", self._on_release)
            self.bind("<Configure>", lambda _event: self._draw())
            self._draw()

        def _parent_bg(self) -> str:
            try:
                return str(self.master.cget("bg"))  # type: ignore[union-attr]
            except Exception:
                return COLORS["app_bg"]

        def _draw(self, bg: Optional[str] = None, border: Optional[str] = None) -> None:
            if not hasattr(self, "canvas"):
                return
            self.canvas.delete("all")
            width = max(2, self.canvas.winfo_width())
            height = max(2, self.canvas.winfo_height())
            disabled = self.state == tk.DISABLED
            fill = bg or (self.disabled_bg if disabled else self.normal_bg)
            outline = "" if disabled else (border if border is not None else self.border)
            text_color = self.disabled_fg if disabled else self.fg
            draw_rounded_rect(
                self.canvas,
                1,
                1,
                width - 2,
                height - 2,
                self.radius,
                fill=fill,
                outline=outline,
                width=1,
            )
            if self.anchor in {"w", "nw", "sw"}:
                x = self.padx + 4
                anchor = tk.W
            else:
                x = width / 2
                anchor = tk.CENTER
            self.canvas.create_text(
                x,
                height / 2,
                text=self.text,
                fill=text_color,
                font=self.font,
                anchor=anchor,
            )
            self.canvas.configure(cursor="arrow" if disabled else "hand2")

        def _on_enter(self, _event: object) -> None:
            if self.state != tk.DISABLED:
                self._draw(self.hover_bg, self.hover_border)

        def _on_leave(self, _event: object) -> None:
            self._pressed = False
            self._draw()

        def _on_press(self, _event: object) -> None:
            if self.state != tk.DISABLED:
                self._pressed = True
                self._draw(self.active_bg, self.active_border)

        def _on_release(self, _event: object) -> None:
            if self.state == tk.DISABLED:
                return
            was_pressed = self._pressed
            self._pressed = False
            self._draw(self.hover_bg, self.hover_border)
            if was_pressed and self.command:
                self.command()

        def configure(self, cnf: Optional[dict[str, object]] = None, **kwargs: object) -> None:  # type: ignore[override]
            options: dict[str, object] = {}
            if cnf:
                options.update(cnf)
            options.update(kwargs)
            frame_options: dict[str, object] = {}
            for key, value in options.items():
                if key == "state":
                    self.state = value  # type: ignore[assignment]
                elif key == "text":
                    self.text = str(value)
                elif key == "command":
                    self.command = value if callable(value) else None
                elif key == "bg":
                    self.normal_bg = str(value)
                    frame_options["bg"] = value
                elif key == "fg":
                    self.fg = str(value)
                elif key == "hover_bg":
                    self.hover_bg = str(value)
                elif key == "active_bg":
                    self.active_bg = str(value)
                elif key == "border":
                    self.border = str(value)
                elif key == "hover_border":
                    self.hover_border = str(value)
                elif key == "active_border":
                    self.active_border = str(value)
                elif key == "font":
                    self.font = value  # type: ignore[assignment]
                else:
                    frame_options[key] = value
            if frame_options:
                super().configure(**frame_options)
            self._draw()

        config = configure

    class CardFrame(tk.Frame):
        def __init__(
            self,
            master: tk.Misc,
            *,
            bg: str = COLORS["card"],
            outer_bg: Optional[str] = None,
            padx: int = 16,
            pady: int = 14,
            radius: int = RADIUS["card"],
            border: str = COLORS["line"],
            shadow: bool = True,
            shadow_color: str = COLORS["shadow"],
        ) -> None:
            super().__init__(master, bg=outer_bg or self._parent_bg(master), highlightthickness=0, bd=0)
            self.card_bg = bg
            self.border = border
            self.radius = radius
            self.padx = padx
            self.pady = pady
            self.shadow = shadow
            self.shadow_color = shadow_color
            self.canvas = tk.Canvas(self, bg=self.cget("bg"), bd=0, highlightthickness=0)
            self.canvas.pack(fill=tk.BOTH, expand=True)
            self.inner = tk.Frame(self.canvas, bg=bg)
            self.inner_window = self.canvas.create_window((padx, pady), window=self.inner, anchor="nw")
            self.canvas.bind("<Configure>", self._on_canvas_configure)
            self.inner.bind("<Configure>", self._on_inner_configure)

        def _parent_bg(self, master: tk.Misc) -> str:
            try:
                return str(master.cget("bg"))  # type: ignore[union-attr]
            except Exception:
                return COLORS["app_bg"]

        def _draw(self) -> None:
            self.canvas.delete("bg")
            width = max(2, self.canvas.winfo_width())
            height = max(2, self.canvas.winfo_height())
            before = set(self.canvas.find_all())
            if self.shadow and width > 12 and height > 12:
                draw_rounded_rect(
                    self.canvas,
                    5,
                    6,
                    width - 2,
                    height - 1,
                    self.radius,
                    fill=self.shadow_color,
                    outline="",
                    width=1,
                )
                card_x2 = width - 5
                card_y2 = height - 6
            else:
                card_x2 = width - 2
                card_y2 = height - 2
            draw_rounded_rect(
                self.canvas,
                1,
                1,
                card_x2,
                card_y2,
                self.radius,
                fill=self.card_bg,
                outline=self.border,
                width=1,
            )
            for item in set(self.canvas.find_all()) - before:
                self.canvas.addtag_withtag("bg", item)
            self.canvas.tag_lower("bg", self.inner_window)

        def _on_canvas_configure(self, event: object) -> None:
            width = int(getattr(event, "width", 200))
            self.canvas.itemconfigure(self.inner_window, width=max(1, width - self.padx * 2 - (4 if self.shadow else 0)))
            self._draw()

        def _on_inner_configure(self, event: object) -> None:
            width = int(getattr(event, "width", 100)) + self.padx * 2 + (4 if self.shadow else 0)
            height = int(getattr(event, "height", 20)) + self.pady * 2 + (6 if self.shadow else 0)
            self.canvas.configure(width=max(1, width), height=max(1, height))


    class PlainCardFrame(tk.Frame):
        """A non-canvas card used for long chat answers.

        The rounded-canvas CardFrame looks better for small controls, but long
        generated answers can be clipped when Tk's requested height is updated
        after wrapping. This plain frame lets content determine its own height,
        so the answer and the citation cards remain fully visible in the main
        scroll area. It intentionally keeps the same .inner/.card_bg/.border
        interface used by MessageBubble.
        """
        def __init__(
            self,
            master: tk.Misc,
            *,
            bg: str = COLORS["card"],
            outer_bg: Optional[str] = None,
            padx: int = 16,
            pady: int = 14,
            radius: int = RADIUS["card"],
            border: str = COLORS["line"],
            shadow: bool = False,
            shadow_color: str = COLORS["shadow"],
        ) -> None:
            super().__init__(
                master,
                bg=bg,
                bd=0,
                highlightthickness=1,
                highlightbackground=border,
                highlightcolor=border,
            )
            self.card_bg = bg
            self.border = border
            self.radius = radius
            self.padx = padx
            self.pady = pady
            self.shadow = shadow
            self.shadow_color = shadow_color
            self.inner = tk.Frame(self, bg=bg, bd=0, highlightthickness=0)
            self.inner.pack(fill=tk.BOTH, expand=True, padx=padx, pady=pady)

        def _draw(self) -> None:
            self.configure(
                bg=self.card_bg,
                highlightbackground=self.border,
                highlightcolor=self.border,
            )
            self.inner.configure(bg=self.card_bg)

    class StatusBar(tk.Frame):
        def __init__(self, master: tk.Misc, app: "App") -> None:
            super().__init__(master, bg=COLORS["sidebar_bg"])
            self.app = app
            self.status = tk.Label(
                self,
                textvariable=app.status_var,
                bg=COLORS["sidebar_bg"],
                fg=COLORS["sidebar_muted"],
                wraplength=200,
                justify=tk.LEFT,
                font=app.font_small,
            )
            self.status.pack(fill=tk.X)
            self.progress = tk.Label(
                self,
                textvariable=app.progress_detail_var,
                bg=COLORS["sidebar_bg"],
                fg="#CBD5E1",
                wraplength=200,
                justify=tk.LEFT,
                font=app.font_small,
            )
            self.progress.pack(fill=tk.X, pady=(8, 0))
            tk.Label(
                self,
                text=f"{DISPLAY_NAME}\n{DISPLAY_SUBTITLE}",
                bg=COLORS["sidebar_bg"],
                fg="#5F728B",
                font=app.font_tiny,
            ).pack(fill=tk.X, pady=(18, 0))

    class Sidebar(tk.Frame):
        def __init__(self, master: tk.Misc, app: "App") -> None:
            super().__init__(master, bg=COLORS["sidebar_bg"], width=244)
            self.app = app
            self.pack_propagate(False)

            brand = tk.Frame(self, bg=COLORS["sidebar_bg"])
            brand.pack(fill=tk.X, padx=20, pady=(24, 18))
            tk.Label(
                brand,
                text=DISPLAY_NAME,
                bg=COLORS["sidebar_bg"],
                fg=COLORS["sidebar_text"],
                justify=tk.LEFT,
                font=app.font_title,
            ).pack(anchor=tk.W)
            tk.Label(
                brand,
                text=DISPLAY_SUBTITLE,
                bg=COLORS["sidebar_bg"],
                fg=COLORS["sidebar_muted"],
                font=app.font_small,
            ).pack(anchor=tk.W, pady=(8, 0))

            state_card = CardFrame(self, bg=COLORS["sidebar_card"], outer_bg=COLORS["sidebar_bg"], radius=RADIUS["card"], border="#1F2A44", shadow=False)
            state_card.pack(fill=tk.X, padx=14, pady=(0, 14))
            tk.Label(
                state_card.inner,
                text="当前资料库",
                bg=COLORS["sidebar_card"],
                fg="#D7DEE9",
                font=app.font_small_bold,
            ).pack(anchor=tk.W, pady=(0, 5))
            self.folder_label = tk.Label(
                state_card.inner,
                text="未选择资料文件夹",
                bg=COLORS["sidebar_card"],
                fg=COLORS["sidebar_muted"],
                wraplength=176,
                justify=tk.LEFT,
                font=app.font_small,
            )
            self.folder_label.pack(anchor=tk.W, pady=(0, 10))
            tk.Label(
                state_card.inner,
                textvariable=app.progress_percent_var,
                bg=COLORS["sidebar_card"],
                fg="#8AB7FF",
                font=app.font_small_bold,
            ).pack(anchor=tk.W)

            ModernButton(
                self,
                "选择资料文件夹",
                app.choose_folder,
                bg="#132035",
                fg="#E2E8F0",
                hover_bg=COLORS["sidebar_hover"],
                active_bg="#24324D",
                border="#24324D",
                hover_border="#334463",
                font=app.font_small_bold,
                anchor="w",
            ).pack(fill=tk.X, padx=14, pady=(0, 8))
            ModernButton(
                self,
                "选择索引库",
                app.choose_db,
                bg="#132035",
                fg="#E2E8F0",
                hover_bg=COLORS["sidebar_hover"],
                active_bg="#24324D",
                border="#24324D",
                hover_border="#334463",
                font=app.font_small_bold,
                anchor="w",
            ).pack(fill=tk.X, padx=14, pady=(0, 8))
            app.index_button = ModernButton(
                self,
                "建立 / 重建索引",
                app.start_index,
                bg=COLORS["primary"],
                fg="#FFFFFF",
                hover_bg=COLORS["primary_dark"],
                active_bg="#1E40AF",
                font=app.font_small_bold,
                anchor="w",
            )
            app.index_button.pack(fill=tk.X, padx=14, pady=(4, 8))
            app.stop_button = ModernButton(
                self,
                "停止索引",
                app.stop_index,
                bg="#1E293B",
                fg="#AAB6C8",
                hover_bg="#27354C",
                active_bg="#334155",
                border="#334155",
                hover_border="#475569",
                font=app.font_small_bold,
                anchor="w",
            )
            app.stop_button.pack(fill=tk.X, padx=14, pady=(0, 20))
            app.stop_button.configure(state=tk.DISABLED)

            recent = tk.Frame(self, bg=COLORS["sidebar_bg"])
            recent.pack(fill=tk.X, padx=14, pady=(2, 0))
            tk.Label(
                recent,
                text="常用问题",
                bg=COLORS["sidebar_bg"],
                fg="#CBD5E1",
                font=app.font_bold,
            ).pack(anchor=tk.W, pady=(0, 8))
            for item in ("制度里怎么规定？", "帮我总结关键要求", "相关文件有哪些？"):
                ModernButton(
                    recent,
                    item,
                    lambda value=item: app.fill_example_question(value),
                    bg=COLORS["sidebar_bg"],
                    fg=COLORS["sidebar_muted"],
                    hover_bg=COLORS["sidebar_hover"],
                    active_bg="#24324D",
                    font=app.font_small,
                    anchor="w",
                    padx=10,
                    pady=6,
                    radius=10,
                ).pack(fill=tk.X, pady=2)

            tk.Frame(self, bg=COLORS["sidebar_bg"]).pack(fill=tk.BOTH, expand=True)
            self.status_bar = StatusBar(self, app)
            self.status_bar.pack(fill=tk.X, padx=14, pady=(0, 16))

        def refresh(self) -> None:
            self.folder_label.configure(text=short_path(self.app.folder_var.get(), 48))

    class CitationCard(CardFrame):
        def __init__(self, master: tk.Misc, app: "App", item: dict[str, str]) -> None:
            super().__init__(master, bg=COLORS["card_alt"], radius=RADIUS["citation"], padx=16, pady=13, border=COLORS["line"], shadow=False)
            self.app = app
            self.item = item
            index = str(item.get("index", ""))
            file_name = str(item.get("file", "未知文件"))
            position = str(item.get("position", ""))
            origin = str(item.get("origin", "本地检索"))
            path = str(item.get("path", ""))
            snippet = str(item.get("snippet", ""))

            header = tk.Frame(self.inner, bg=COLORS["card_alt"])
            header.pack(fill=tk.X)
            title = tk.Frame(header, bg=COLORS["card_alt"])
            title.pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(
                title,
                text=f"[{index}] {file_name}" if index else file_name,
                bg=COLORS["card_alt"],
                fg=COLORS["text"],
                font=app.font_bold,
                anchor=tk.W,
            ).pack(fill=tk.X)
            if position or origin:
                tk.Label(
                    title,
                    text=" · ".join(part for part in (position, origin) if part),
                    bg=COLORS["card_alt"],
                    fg=COLORS["muted"],
                    font=app.font_small,
                    anchor=tk.W,
                ).pack(fill=tk.X, pady=(4, 0))

            actions = tk.Frame(header, bg=COLORS["card_alt"])
            actions.pack(side=tk.RIGHT, padx=(12, 0))
            ModernButton(
                actions,
                "打开文件",
                lambda: app.open_citation(item),
                bg=COLORS["primary_soft"],
                fg=COLORS["primary"],
                hover_bg="#DBEAFE",
                active_bg="#CFE0FF",
                border="#D6E6FF",
                font=app.font_small_bold,
                padx=10,
                pady=5,
                radius=10,
                min_width=76,
            ).pack(side=tk.LEFT)
            if path:
                ModernButton(
                    actions,
                    "复制路径",
                    lambda value=path: app.copy_to_clipboard(value, "文件路径"),
                    bg=COLORS["card"],
                    fg=COLORS["muted"],
                    hover_bg="#F1F5F9",
                    active_bg="#E2E8F0",
                    border=COLORS["line"],
                    font=app.font_small,
                    padx=9,
                    pady=5,
                    radius=10,
                    min_width=76,
                ).pack(side=tk.LEFT, padx=(8, 0))

            if snippet:
                snippet_text = tk.Text(
                    self.inner,
                    height=1,
                    wrap=tk.WORD,
                    relief=tk.FLAT,
                    bd=0,
                    padx=0,
                    pady=0,
                    bg=COLORS["card_alt"],
                    fg="#334155",
                    insertbackground=COLORS["primary"],
                    font=app.font_small,
                    cursor="xterm",
                )
                snippet_text.pack(fill=tk.X, pady=(10, 4))
                snippet_text.insert("1.0", snippet)
                snippet_text.configure(state=tk.DISABLED)
                snippet_lines = max(2, min(5, 1 + len(snippet) // 70 + snippet.count("\n")))
                snippet_text.configure(height=snippet_lines)

            if path:
                path_frame = tk.Frame(self.inner, bg=COLORS["card_alt"])
                path_frame.pack(fill=tk.X, pady=(4, 0))
                tk.Label(
                    path_frame,
                    text=short_path(path, 98),
                    bg=COLORS["card_alt"],
                    fg="#94A3B8",
                    font=app.font_tiny,
                    justify=tk.LEFT,
                    anchor=tk.W,
                ).pack(side=tk.LEFT, fill=tk.X, expand=True)

    class MessageBubble(tk.Frame):
        def __init__(
            self,
            master: tk.Misc,
            app: "App",
            role: str,
            text: str,
            *,
            turn_no: int = 0,
            citations: Optional[list[dict[str, str]]] = None,
            pending: bool = False,
            error: bool = False,
        ) -> None:
            super().__init__(master, bg=COLORS["app_bg"])
            self.app = app
            self.role = role
            self.turn_no = turn_no
            self.pending = pending
            self.error = error
            self.citations: list[dict[str, str]] = citations or []
            self.raw_text = app.clean_chat_markdown(text) if role == "assistant" else text.strip()
            self._resize_body_job: Optional[str] = None
            self._last_body_width = 0
            align_right = role == "user"

            outer = tk.Frame(self, bg=COLORS["app_bg"])
            outer.pack(fill=tk.X, pady=(8 if align_right else 12, 10 if align_right else 14))
            bubble_bg, border = self._style_colors()
            self.card = PlainCardFrame(
                outer,
                bg=bubble_bg,
                radius=RADIUS["bubble"],
                padx=(18 if align_right else 20),
                pady=(12 if align_right else 16),
                border=border,
                shadow=False,
            )
            if align_right:
                self.card.pack(side=tk.RIGHT, padx=(340, 56), fill=tk.X, expand=True)
            else:
                self.card.pack(side=tk.LEFT, padx=(56, 72), fill=tk.X, expand=True)

            header = tk.Frame(self.card.inner, bg=bubble_bg)
            header.pack(fill=tk.X, pady=(0, 8))
            self.title_label = tk.Label(
                header,
                text="",
                bg=bubble_bg,
                fg=COLORS["primary"] if align_right else COLORS["success"],
                font=app.font_bold,
                anchor=tk.W,
            )
            self.title_label.pack(side=tk.LEFT, fill=tk.X, expand=True)

            self.action_bar = tk.Frame(header, bg=bubble_bg)
            if role == "assistant":
                self.action_bar.pack(side=tk.RIGHT, padx=(12, 0))
                ModernButton(
                    self.action_bar,
                    "复制回答",
                    lambda: app.copy_to_clipboard(self.raw_text, "回答内容"),
                    bg=COLORS["card_alt"],
                    fg=COLORS["muted"],
                    hover_bg="#EEF2FF",
                    active_bg="#E0E7FF",
                    border=COLORS["line"],
                    font=app.font_small,
                    padx=10,
                    pady=4,
                    radius=10,
                    min_width=72,
                ).pack(side=tk.LEFT)
                ModernButton(
                    self.action_bar,
                    "复制含引用",
                    lambda: app.copy_to_clipboard(app.compose_answer_with_citations(self.raw_text, self.citations), "回答和引用"),
                    bg=COLORS["card_alt"],
                    fg=COLORS["muted"],
                    hover_bg="#EEF2FF",
                    active_bg="#E0E7FF",
                    border=COLORS["line"],
                    font=app.font_small,
                    padx=10,
                    pady=4,
                    radius=10,
                    min_width=86,
                ).pack(side=tk.LEFT, padx=(8, 0))

            # 用 Label 展示回答正文，避免 Text 控件高度计算不准导致内容被截断。
            # 复制能力由右上角按钮和右键菜单提供；对话区整体负责滚动。
            self.body_text = tk.Label(
                self.card.inner,
                text=self.raw_text,
                bg=bubble_bg,
                fg=COLORS["text"] if not error else "#991B1B",
                font=app.font_chat,
                justify=tk.LEFT,
                anchor=tk.NW,
                wraplength=820,
            )
            self.body_text.pack(fill=tk.X, expand=False)
            self.body_text.bind("<Button-3>", self._show_copy_menu)
            self.body_text.bind("<Control-c>", self._copy_selected)
            self.body_text.bind("<MouseWheel>", self._forward_mousewheel)
            self.body_text.bind("<Button-4>", self._forward_mousewheel)
            self.body_text.bind("<Button-5>", self._forward_mousewheel)
            self._set_body_text(self.raw_text)

            self.citation_container = tk.Frame(self.card.inner, bg=bubble_bg)
            self.set_citations(self.citations)
            self._update_title()
            self.card.inner.bind("<Configure>", self._on_inner_configure)

        def _forward_mousewheel(self, event: object) -> str:
            if hasattr(self.app, "chat_area"):
                return self.app.chat_area._on_mousewheel(event)
            return "break"

        def _show_copy_menu(self, event: object) -> str:
            menu = tk.Menu(self, tearoff=0)
            menu.add_command(label="复制选中内容", command=self._copy_selected)
            menu.add_command(label="复制整段回答", command=lambda: self.app.copy_to_clipboard(self.raw_text, "整段内容"))
            try:
                menu.tk_popup(int(getattr(event, "x_root", 0)), int(getattr(event, "y_root", 0)))
            finally:
                menu.grab_release()
            return "break"

        def _copy_selected(self, _event: object = None) -> str:
            self.app.copy_to_clipboard(self.raw_text, "回答内容")
            return "break"

        def _set_body_text(self, body: str) -> None:
            self.body_text.configure(text=body)
            self._schedule_body_resize()

        def _schedule_body_resize(self, delay: int = 25) -> None:
            if self._resize_body_job:
                try:
                    self.after_cancel(self._resize_body_job)
                except Exception:
                    pass
            self._resize_body_job = self.after(delay, self._resize_body_text)

        def _resize_body_text(self) -> None:
            self._resize_body_job = None
            try:
                width = max(320, self.card.inner.winfo_width())
            except Exception:
                width = 760
            # Label 自然撑开高度，只需要根据当前宽度更新换行宽度，不再设置 height，避免截断。
            wrap = max(320, width - 8)
            self.body_text.configure(wraplength=wrap)
            if hasattr(self.app, "chat_area"):
                self.app.chat_area.request_scrollregion_refresh(20)
                self.app.chat_area.request_scrollregion_refresh(160)

        def _on_inner_configure(self, event: object) -> None:
            width = int(getattr(event, "width", 760))
            if abs(width - self._last_body_width) < 16:
                return
            self._last_body_width = width
            self._schedule_body_resize()

        def _style_colors(self) -> tuple[str, str]:
            if self.error:
                return "#FEF2F2", "#FECACA"
            if self.role == "user":
                return COLORS["bubble_user"], "#CFE0FF"
            return COLORS["card"], COLORS["line"]

        def _update_title(self) -> None:
            if self.role == "user":
                title = f"你 · 第 {self.turn_no} 轮" if self.turn_no else "你"
            elif self.pending:
                title = f"{DISPLAY_NAME} · 正在处理"
            else:
                title = DISPLAY_NAME
            self.title_label.configure(text=title)

        def _apply_style(self) -> None:
            bubble_bg, border = self._style_colors()
            self.card.card_bg = bubble_bg
            self.card.border = border
            self.card.inner.configure(bg=bubble_bg)
            self.title_label.configure(bg=bubble_bg)
            self.action_bar.configure(bg=bubble_bg)
            self.body_text.configure(bg=bubble_bg, fg=COLORS["text"] if not self.error else "#991B1B")
            self.citation_container.configure(bg=bubble_bg)
            self.card._draw()

        def update_text(self, text: str, *, pending: Optional[bool] = None, error: Optional[bool] = None) -> None:
            if pending is not None:
                self.pending = pending
            if error is not None:
                self.error = error
            self.raw_text = self.app.clean_chat_markdown(text) if self.role == "assistant" else text.strip()
            self._set_body_text(self.raw_text)
            self._update_title()
            self._apply_style()

        def set_citations(self, citations: Optional[list[dict[str, str]]]) -> None:
            self.citations = citations or []
            for child in self.citation_container.winfo_children():
                child.destroy()
            if not self.citations:
                self.citation_container.pack_forget()
                return
            bubble_bg, _border = self._style_colors()
            self.citation_container.configure(bg=bubble_bg)
            self.citation_container.pack(fill=tk.X, pady=(16, 0))
            header = tk.Frame(self.citation_container, bg=bubble_bg)
            header.pack(fill=tk.X, pady=(0, 9))
            tk.Label(
                header,
                text=f"相关索引文件 · {len(self.citations)} 条来源",
                bg=bubble_bg,
                fg=COLORS["muted"],
                font=self.app.font_bold,
                anchor=tk.W,
            ).pack(side=tk.LEFT, fill=tk.X, expand=True)
            ModernButton(
                header,
                "复制引用",
                lambda: self.app.copy_to_clipboard(self.app.compose_citation_text(self.citations), "引用资料"),
                bg=COLORS["card_alt"],
                fg=COLORS["muted"],
                hover_bg="#EEF2FF",
                active_bg="#E0E7FF",
                border=COLORS["line"],
                font=self.app.font_small,
                padx=10,
                pady=4,
                radius=10,
                min_width=76,
            ).pack(side=tk.RIGHT)
            for citation in self.citations[:10]:
                card = CitationCard(self.citation_container, self.app, citation)
                card.pack(fill=tk.X, pady=(0, 8))
            if len(self.citations) > 10:
                tk.Label(
                    self.citation_container,
                    text=f"还有 {len(self.citations) - 10} 条引用已用于回答，可通过“复制引用”查看完整来源。",
                    bg=bubble_bg,
                    fg=COLORS["muted"],
                    font=self.app.font_small,
                    anchor=tk.W,
                ).pack(fill=tk.X, pady=(0, 4))
            if hasattr(self.app, "chat_area"):
                self.app.chat_area.request_scrollregion_refresh()

    class ChatArea(tk.Frame):
        def __init__(self, master: tk.Misc, app: "App") -> None:
            super().__init__(master, bg=COLORS["app_bg"])
            self.app = app
            self.columnconfigure(0, weight=1)
            self.rowconfigure(1, weight=1)
            self.rendered_history_count = 0
            self.pending_question = ""
            self.pending_user_bubble: Optional[MessageBubble] = None
            self.pending_assistant_bubble: Optional[MessageBubble] = None
            self.welcome_widget: Optional[tk.Widget] = None
            self.user_scroll_locked = False
            self._scrollregion_job: Optional[str] = None
            self._resize_job: Optional[str] = None

            top = tk.Frame(self, bg=COLORS["app_bg"])
            top.grid(row=0, column=0, sticky=tk.EW, padx=42, pady=(24, 8))
            title_box = tk.Frame(top, bg=COLORS["app_bg"])
            title_box.pack(side=tk.LEFT, fill=tk.X, expand=True)
            tk.Label(
                title_box,
                text="当前对话",
                bg=COLORS["app_bg"],
                fg=COLORS["text"],
                font=app.font_heading,
            ).pack(anchor=tk.W)
            tk.Label(
                title_box,
                text="本地检索优先 · 回答后紧跟相关索引文件 · 可自由滚动查看",
                bg=COLORS["app_bg"],
                fg=COLORS["muted"],
                font=app.font_small,
            ).pack(anchor=tk.W, pady=(3, 0))
            self.latest_button = ModernButton(
                top,
                "回到最新",
                self.scroll_to_bottom,
                bg=COLORS["primary_soft"],
                fg=COLORS["primary"],
                hover_bg="#DBEAFE",
                active_bg="#CFE0FF",
                border="#D6E6FF",
                font=app.font_small_bold,
                padx=14,
                pady=6,
                radius=13,
                min_width=88,
            )
            self.latest_button.pack(side=tk.RIGHT, padx=(10, 0))
            self.clear_button = ModernButton(
                top,
                "清空对话",
                app.clear_rag_history,
                bg=COLORS["card"],
                fg=COLORS["text"],
                hover_bg="#F0F5FF",
                active_bg="#E7F0FF",
                border=COLORS["line"],
                font=app.font_small_bold,
                padx=14,
                pady=6,
                radius=13,
                min_width=88,
            )
            self.clear_button.pack(side=tk.RIGHT, padx=(10, 0))
            self.settings_button = ModernButton(
                top,
                "模型与检索",
                app.toggle_settings_panel,
                bg=COLORS["card"],
                fg=COLORS["text"],
                hover_bg="#F0F5FF",
                active_bg="#E7F0FF",
                border=COLORS["line"],
                font=app.font_small_bold,
                padx=14,
                pady=6,
                radius=13,
                min_width=100,
            )
            self.settings_button.pack(side=tk.RIGHT, padx=(10, 0))
            status_pill = tk.Frame(top, bg=COLORS["app_bg"])
            status_pill.pack(side=tk.RIGHT)
            self.header_status = tk.Label(
                status_pill,
                textvariable=app.rag_status_var,
                bg=COLORS["app_bg"],
                fg=COLORS["muted"],
                font=app.font_small,
            )
            self.header_status.pack(side=tk.RIGHT)

            self.canvas = tk.Canvas(self, bg=COLORS["app_bg"], bd=0, highlightthickness=0, yscrollincrement=18)
            self.scrollbar = tk.Scrollbar(
                self,
                orient=tk.VERTICAL,
                command=self._on_scrollbar,
                width=24,
                bd=0,
                relief=tk.FLAT,
                bg="#7C8EA6",
                activebackground="#5B6F89",
                troughcolor="#D8E1EE",
                highlightthickness=0,
                elementborderwidth=0,
            )
            self.scroll_frame = tk.Frame(self.canvas, bg=COLORS["app_bg"])
            self.scroll_window = self.canvas.create_window((0, 0), window=self.scroll_frame, anchor="nw")
            self.canvas.configure(yscrollcommand=self._on_yview)
            self.canvas.grid(row=1, column=0, sticky=tk.NSEW, padx=(0, 0))
            self.scrollbar.grid(row=1, column=1, sticky=tk.NS, padx=(6, 18), pady=(12, 12))
            self.scroll_frame.bind("<Configure>", self._on_frame_configure)
            self.canvas.bind("<Configure>", self._on_canvas_configure)
            self.canvas.bind("<Enter>", self._bind_mousewheel)
            self.canvas.bind("<Leave>", self._unbind_mousewheel)
            self.scroll_frame.bind("<Enter>", self._bind_mousewheel)

            input_shell = tk.Frame(self, bg=COLORS["app_bg"])
            input_shell.grid(row=2, column=0, columnspan=2, sticky=tk.EW, padx=64, pady=(10, 24))
            input_card = CardFrame(input_shell, bg=COLORS["card"], radius=RADIUS["input"], padx=12, pady=10, border=COLORS["line"], shadow=True)
            input_card.pack(fill=tk.X)
            input_card.inner.columnconfigure(0, weight=1)
            app.chat_input = tk.Text(
                input_card.inner,
                height=3,
                wrap=tk.WORD,
                relief=tk.FLAT,
                bd=0,
                padx=16,
                pady=12,
                bg=COLORS["card"],
                fg=COLORS["text"],
                insertbackground=COLORS["primary"],
                font=app.font_input,
            )
            app.chat_input.grid(row=0, column=0, sticky=tk.EW)
            app.chat_input.bind("<Return>", app.handle_chat_return)
            app.chat_input.bind("<Shift-Return>", lambda _event: None)
            app.rag_button = ModernButton(
                input_card.inner,
                "发送",
                app.run_rag_answer,
                bg=COLORS["primary"],
                fg="#FFFFFF",
                hover_bg=COLORS["primary_dark"],
                active_bg="#1E40AF",
                font=app.font_small_bold,
                padx=18,
                pady=10,
                radius=16,
                min_width=78,
            )
            app.rag_button.grid(row=0, column=1, sticky=tk.NS, padx=(10, 8), pady=8)
            tk.Label(
                input_shell,
                text=SECURITY_HINT,
                bg=COLORS["app_bg"],
                fg=COLORS["muted"],
                font=app.font_tiny,
                anchor=tk.W,
                justify=tk.LEFT,
                wraplength=940,
            ).pack(fill=tk.X, pady=(8, 0))

        def request_scrollregion_refresh(self, delay: int = 35) -> None:
            if self._scrollregion_job:
                try:
                    self.after_cancel(self._scrollregion_job)
                except Exception:
                    pass
            self._scrollregion_job = self.after(delay, self._refresh_scrollregion)

        def _on_frame_configure(self, _event: object) -> None:
            self.request_scrollregion_refresh()

        def _on_canvas_configure(self, event: object) -> None:
            width = int(getattr(event, "width", 800))
            if self._resize_job:
                self.after_cancel(self._resize_job)
            self._resize_job = self.after(70, lambda value=width: self._resize_scroll_window(value))

        def _refresh_scrollregion(self) -> None:
            self._scrollregion_job = None
            try:
                self.update_idletasks()
            except Exception:
                pass
            bbox = self.canvas.bbox("all")
            if bbox:
                # 留出充足的底部安全区，避免最后一行回答被输入框视觉上遮住。
                self.canvas.configure(scrollregion=(bbox[0], bbox[1], bbox[2], bbox[3] + 140))

        def _resize_scroll_window(self, width: int) -> None:
            self._resize_job = None
            current = int(float(self.canvas.itemcget(self.scroll_window, "width") or 0))
            if abs(current - width) > 2:
                self.canvas.itemconfigure(self.scroll_window, width=width)

        def _on_scrollbar(self, *args: str) -> None:
            self.user_scroll_locked = True
            self.canvas.yview(*args)

        def _on_yview(self, first: str, last: str) -> None:
            self.scrollbar.set(first, last)
            try:
                if float(last) >= 0.985:
                    self.user_scroll_locked = False
            except ValueError:
                pass

        def _bind_mousewheel(self, _event: object) -> None:
            self.canvas.bind_all("<MouseWheel>", self._on_mousewheel)
            self.canvas.bind_all("<Button-4>", self._on_mousewheel)
            self.canvas.bind_all("<Button-5>", self._on_mousewheel)

        def _unbind_mousewheel(self, _event: object) -> None:
            self.canvas.unbind_all("<MouseWheel>")
            self.canvas.unbind_all("<Button-4>")
            self.canvas.unbind_all("<Button-5>")

        def _on_mousewheel(self, event: object) -> str:
            self.user_scroll_locked = True
            delta = int(getattr(event, "delta", 0))
            num = int(getattr(event, "num", 0))
            if num == 4:
                units = -2
            elif num == 5:
                units = 2
            else:
                steps = int(delta / 120)
                if steps == 0 and delta:
                    steps = 1 if delta > 0 else -1
                units = -1 * max(-3, min(3, steps)) * 2
            self.canvas.yview_scroll(units, "units")
            return "break"

        def clear_messages(self) -> None:
            for child in self.scroll_frame.winfo_children():
                child.destroy()
            self.rendered_history_count = 0
            self.pending_question = ""
            self.pending_user_bubble = None
            self.pending_assistant_bubble = None
            self.welcome_widget = None

        def show_welcome(self) -> None:
            if self.welcome_widget and self.welcome_widget.winfo_exists():
                return
            self.clear_messages()
            wrap = tk.Frame(self.scroll_frame, bg=COLORS["app_bg"])
            self.welcome_widget = wrap
            wrap.pack(fill=tk.BOTH, expand=True, padx=42, pady=58)

            head = tk.Frame(wrap, bg=COLORS["app_bg"])
            head.pack(fill=tk.X, padx=26)
            tk.Label(
                head,
                text="今天要查什么资料？",
                bg=COLORS["app_bg"],
                fg=COLORS["text"],
                font=self.app.font_hero,
            ).pack(anchor=tk.W, pady=(0, 10))
            tk.Label(
                head,
                text=f"选择资料库并建立索引后，直接输入问题。系统会先检索本地资料，再生成带引用的回答。{AI_REVIEW_HINT}",
                bg=COLORS["app_bg"],
                fg=COLORS["muted"],
                font=self.app.font_chat,
                wraplength=900,
                justify=tk.LEFT,
            ).pack(anchor=tk.W, pady=(0, 26))

            grid = tk.Frame(wrap, bg=COLORS["app_bg"])
            grid.pack(fill=tk.X, padx=26)
            examples = (
                ("查制度", "制度里怎么规定？"),
                ("提炼要点", "帮我总结关键要求"),
                ("找文件", "相关文件有哪些？"),
                ("转成表格", "把检索结果整理成表格"),
            )
            for i, (title, prompt) in enumerate(examples):
                tile = CardFrame(grid, bg=COLORS["card"], radius=16, padx=18, pady=16, border=COLORS["line"], shadow=False)
                tile.grid(row=i // 2, column=i % 2, sticky=tk.NSEW, padx=(0 if i % 2 == 0 else 10, 10 if i % 2 == 0 else 0), pady=6)
                grid.columnconfigure(i % 2, weight=1)
                tk.Label(tile.inner, text=title, bg=COLORS["card"], fg=COLORS["text"], font=self.app.font_small_bold).pack(anchor=tk.W)
                tk.Label(tile.inner, text=prompt, bg=COLORS["card"], fg=COLORS["muted"], font=self.app.font_small, wraplength=360, justify=tk.LEFT).pack(anchor=tk.W, pady=(6, 0))
                tile.inner.bind("<Button-1>", lambda _e, value=prompt: self.app.fill_example_question(value))
                for child in tile.inner.winfo_children():
                    child.bind("<Button-1>", lambda _e, value=prompt: self.app.fill_example_question(value))

            actions = tk.Frame(wrap, bg=COLORS["app_bg"])
            actions.pack(fill=tk.X, padx=26, pady=(18, 0))
            ModernButton(
                actions,
                "选择资料文件夹",
                self.app.choose_folder,
                bg=COLORS["primary"],
                fg="#FFFFFF",
                hover_bg=COLORS["primary_dark"],
                active_bg="#1E40AF",
                font=self.app.font_bold,
                min_width=136,
            ).pack(side=tk.LEFT)
            ModernButton(
                actions,
                "打开模型设置",
                self.app.toggle_settings_panel,
                bg=COLORS["card"],
                fg=COLORS["text"],
                hover_bg="#F0F5FF",
                active_bg="#E7F0FF",
                border=COLORS["line"],
                font=self.app.font_bold,
                min_width=126,
            ).pack(side=tk.LEFT, padx=(10, 0))
            self.canvas.after(80, lambda: self.canvas.yview_moveto(0))

        def hide_welcome(self) -> None:
            if self.welcome_widget and self.welcome_widget.winfo_exists():
                self.welcome_widget.destroy()
            self.welcome_widget = None

        def add_message(
            self,
            role: str,
            text: str,
            *,
            turn_no: int = 0,
            citations: Optional[list[dict[str, str]]] = None,
            pending: bool = False,
            error: bool = False,
            scroll: str = "none",
        ) -> MessageBubble:
            bubble = MessageBubble(
                self.scroll_frame,
                self.app,
                role,
                text,
                turn_no=turn_no,
                citations=citations,
                pending=pending,
                error=error,
            )
            bubble.pack(fill=tk.X)
            bubble.bind("<Enter>", self._bind_mousewheel)
            self.request_scrollregion_refresh()
            if scroll == "bottom":
                self.canvas.after(90, self.scroll_to_bottom)
            elif scroll == "top":
                self.canvas.after(80, lambda widget=bubble: self.scroll_to_widget_top(widget))
                self.canvas.after(220, lambda widget=bubble: self.scroll_to_widget_top(widget))
            return bubble

        def _append_history_from(self, history: list[dict[str, object]], start: int) -> None:
            turn_no = start // 2 + 1
            index = start
            while index + 1 < len(history):
                user_item = history[index]
                assistant_item = history[index + 1]
                self.add_message("user", str(user_item.get("content", "")), turn_no=turn_no)
                self.add_message(
                    "assistant",
                    str(assistant_item.get("content", "")),
                    citations=assistant_item.get("source_items", []) if isinstance(assistant_item, dict) else [],
                )
                self.rendered_history_count = index + 2
                turn_no += 1
                index += 2

        def _finalize_pending_from_history(self, history: list[dict[str, object]]) -> bool:
            if not self.pending_question or not self.pending_assistant_bubble:
                return False
            if len(history) < self.rendered_history_count + 2:
                return False
            user_item = history[self.rendered_history_count]
            assistant_item = history[self.rendered_history_count + 1]
            if str(user_item.get("content", "")) != self.pending_question:
                return False
            self.pending_assistant_bubble.update_text(str(assistant_item.get("content", "")), pending=False, error=False)
            if isinstance(assistant_item, dict):
                self.pending_assistant_bubble.set_citations(assistant_item.get("source_items", []))
            self.rendered_history_count += 2
            answer_bubble = self.pending_assistant_bubble
            self.pending_question = ""
            self.pending_user_bubble = None
            self.pending_assistant_bubble = None
            if not self.user_scroll_locked:
                self.canvas.after(90, lambda widget=answer_bubble: self.scroll_to_widget_top(widget))
                self.canvas.after(260, lambda widget=answer_bubble: self.scroll_to_widget_top(widget))
            return True

        def _sync_history(self, history: list[dict[str, object]]) -> None:
            if not history:
                if self.rendered_history_count:
                    self.clear_messages()
                return
            self.hide_welcome()
            if len(history) < self.rendered_history_count:
                self.clear_messages()
            finalized = self._finalize_pending_from_history(history)
            start = self.rendered_history_count
            if finalized:
                start = self.rendered_history_count
            if start < len(history):
                self._append_history_from(history, start)

        def start_or_update_pending(self, question: str, answer: str, *, error: bool = False) -> None:
            self.hide_welcome()
            if self.pending_question == question and self.pending_assistant_bubble:
                self.pending_assistant_bubble.update_text(answer or "正在检索本地知识库…\n正在生成回答…", pending=not error, error=error)
                return
            if self.pending_question and self.pending_assistant_bubble:
                self.pending_assistant_bubble.update_text(
                    self.pending_assistant_bubble.raw_text,
                    pending=False,
                    error=self.pending_assistant_bubble.error,
                )
            self.pending_question = question
            self.user_scroll_locked = False
            turn_no = self.rendered_history_count // 2 + 1
            self.pending_user_bubble = self.add_message("user", question, turn_no=turn_no)
            self.pending_assistant_bubble = self.add_message(
                "assistant",
                answer or "正在检索本地知识库…\n正在生成回答…",
                pending=not error,
                error=error,
                scroll="top",
            )

        def scroll_to_widget_top(self, widget: tk.Widget) -> None:
            self.update_idletasks()
            bbox = self.canvas.bbox("all")
            if not bbox:
                return
            padded_bbox = (bbox[0], bbox[1], bbox[2], bbox[3] + 28)
            self.canvas.configure(scrollregion=padded_bbox)
            content_height = max(1, padded_bbox[3] - padded_bbox[1])
            canvas_height = max(1, self.canvas.winfo_height())
            max_fraction = max(0.0, 1.0 - (canvas_height / content_height))
            y = max(0, widget.winfo_y() - 18)
            target = max(0.0, min(max_fraction, y / content_height))
            self.canvas.yview_moveto(target)

        def scroll_to_bottom(self) -> None:
            self.user_scroll_locked = False
            try:
                self.update_idletasks()
            except Exception:
                pass
            bbox = self.canvas.bbox("all")
            if bbox:
                self.canvas.configure(scrollregion=(bbox[0], bbox[1], bbox[2], bbox[3] + 28))
            self.canvas.yview_moveto(1.0)

        def render(
            self,
            history: list[dict[str, object]],
            *,
            pending_question: str = "",
            pending_answer: str = "",
            error: bool = False,
        ) -> None:
            if not history and not pending_question:
                self.show_welcome()
                return
            self._sync_history(history)
            if pending_question:
                self.start_or_update_pending(pending_question, pending_answer, error=error)

    class SettingsPanel(tk.Frame):
        def __init__(self, master: tk.Misc, app: "App") -> None:
            super().__init__(master, bg=COLORS["app_bg"], width=318)
            self.app = app
            self.pack_propagate(False)

            tk.Label(
                self,
                text="模型与检索",
                bg=COLORS["app_bg"],
                fg=COLORS["text"],
                font=app.font_heading,
            ).pack(anchor=tk.W, padx=20, pady=(24, 12))

            model_card = CardFrame(self, radius=RADIUS["card"], padx=18, pady=18, border=COLORS["line"])
            model_card.pack(fill=tk.X, padx=18, pady=(0, 14))
            tk.Label(model_card.inner, text="模型设置", bg=COLORS["card"], fg=COLORS["text"], font=app.font_heading).pack(anchor=tk.W)
            self._entry(model_card.inner, "API 地址", app.rag_api_url_var)
            self._entry(model_card.inner, "API Key", app.rag_api_key_var, show="*")
            tk.Label(model_card.inner, text=f"{AI_REVIEW_HINT} 请勿上传或输入涉密、敏感、客户隐私、未公开经营数据。", bg=COLORS["card"], fg=COLORS["muted"], wraplength=270, justify=tk.LEFT, font=app.font_tiny).pack(anchor=tk.W, pady=(5, 9))
            tk.Label(model_card.inner, text="模型", bg=COLORS["card"], fg=COLORS["muted"], font=app.font_small).pack(anchor=tk.W, pady=(4, 3))
            combo = ttk.Combobox(
                model_card.inner,
                textvariable=app.rag_model_var,
                values=("deepseek-chat", "deepseek-reasoner", "deepseek-v4-pro", "gpt-4o", "gpt-4o-mini"),
                width=27,
            )
            combo.pack(fill=tk.X, ipady=4)
            tk.Checkbutton(
                model_card.inner,
                text="保存 Key 到本机",
                variable=app.rag_save_key_var,
                bg=COLORS["card"],
                fg=COLORS["text"],
                activebackground=COLORS["card"],
                selectcolor=COLORS["card"],
                font=app.font_small,
                bd=0,
            ).pack(anchor=tk.W, pady=(10, 4))
            ModernButton(model_card.inner, "保存配置", app.save_current_rag_settings, bg=COLORS["primary_soft"], fg=COLORS["primary"], hover_bg="#DBEAFE", active_bg="#CFE0FF", border="#D6E6FF", font=app.font_small_bold).pack(fill=tk.X, pady=(8, 0))

            retrieval_card = CardFrame(self, radius=RADIUS["card"], padx=18, pady=18, border=COLORS["line"])
            retrieval_card.pack(fill=tk.X, padx=18, pady=(0, 14))
            tk.Label(retrieval_card.inner, text="检索设置", bg=COLORS["card"], fg=COLORS["text"], font=app.font_heading).pack(anchor=tk.W)
            row = tk.Frame(retrieval_card.inner, bg=COLORS["card"])
            row.pack(fill=tk.X, pady=(8, 6))
            tk.Label(row, text="引用片段", bg=COLORS["card"], fg=COLORS["muted"], font=app.font_small).pack(side=tk.LEFT)
            ttk.Spinbox(row, from_=1, to=20, width=6, textvariable=app.rag_top_k_var).pack(side=tk.RIGHT)
            self._check(retrieval_card.inner, "按完整短语检索", app.phrase_var)
            self._check(retrieval_card.inner, "同一文本块包含全部关键词", app.require_all_var)
            self._check(retrieval_card.inner, "区分大小写", app.match_case_var)

        def _entry(self, master: tk.Misc, label: str, var: tk.StringVar, show: str = "") -> None:
            tk.Label(master, text=label, bg=COLORS["card"], fg=COLORS["muted"], font=self.app.font_small).pack(anchor=tk.W, pady=(10, 3))
            tk.Entry(
                master,
                textvariable=var,
                show=show,
                bg=COLORS["card_alt"],
                fg=COLORS["text"],
                relief=tk.FLAT,
                bd=0,
                highlightthickness=1,
                highlightbackground=COLORS["line"],
                highlightcolor=COLORS["primary"],
                insertbackground=COLORS["primary"],
                font=self.app.font_ui,
            ).pack(fill=tk.X, ipady=7)

        def _check(self, master: tk.Misc, label: str, var: tk.BooleanVar) -> None:
            tk.Checkbutton(
                master,
                text=label,
                variable=var,
                bg=COLORS["card"],
                fg=COLORS["text"],
                activebackground=COLORS["card"],
                selectcolor=COLORS["card"],
                font=self.app.font_small,
                bd=0,
            ).pack(anchor=tk.W, pady=3)

    class App(tk.Tk):
        def __init__(self) -> None:
            super().__init__()
            self.title(DISPLAY_NAME)
            icon_path = resource_path("app_icon.ico")
            if icon_path.exists():
                try:
                    self.iconbitmap(default=str(icon_path))
                except Exception:
                    pass
            self.geometry("1440x860")
            self.minsize(1180, 740)
            self.configure(bg=COLORS["app_bg"])

            self.folder_var = tk.StringVar()
            self.db_var = tk.StringVar(value=str(default_db_path()))
            self.query_var = tk.StringVar()
            self.max_mb_var = tk.IntVar(value=DEFAULT_MAX_FILE_MB)
            self.limit_var = tk.IntVar(value=500)
            self.require_all_var = tk.BooleanVar(value=False)
            self.match_case_var = tk.BooleanVar(value=False)
            self.phrase_var = tk.BooleanVar(value=False)
            rag_settings = load_rag_settings()
            self.rag_api_url_var = tk.StringVar(value=str(rag_settings.get("api_url", DEFAULT_RAG_API_URL)))
            self.rag_model_var = tk.StringVar(value=str(rag_settings.get("model", DEFAULT_RAG_MODEL)))
            self.rag_api_key_var = tk.StringVar(value=str(rag_settings.get("api_key", "")))
            self.rag_top_k_var = tk.IntVar(value=int(rag_settings.get("top_k", DEFAULT_RAG_TOP_K) or DEFAULT_RAG_TOP_K))
            self.rag_save_key_var = tk.BooleanVar(value=bool(rag_settings.get("save_key", False)))
            self.rag_status_var = tk.StringVar(value=f"{DISPLAY_NAME}就绪")
            self.status_var = tk.StringVar(value=f"{DISPLAY_NAME}就绪：请选择文件夹并建立索引")
            self.progress_var = tk.DoubleVar(value=0.0)
            self.progress_percent_var = tk.StringVar(value="0%")
            self.progress_detail_var = tk.StringVar(value="就绪")
            self.queue: "queue.Queue[dict[str, object]]" = queue.Queue(maxsize=200)
            self.last_hits: list[SearchHit] = []
            self.rag_history: list[dict[str, str]] = []
            self.worker: Optional[threading.Thread] = None
            self.rag_worker: Optional[threading.Thread] = None
            self.rag_cancel_requested = False
            self.stop_event = threading.Event()

            self.create_widgets()
            self.after(150, self.drain_queue)

        def create_widgets(self) -> None:
            self.font_ui = tkfont.Font(family="Microsoft YaHei UI", size=11)
            self.font_small = tkfont.Font(family="Microsoft YaHei UI", size=10)
            self.font_tiny = tkfont.Font(family="Microsoft YaHei UI", size=9)
            self.font_bold = tkfont.Font(family="Microsoft YaHei UI", size=11, weight="bold")
            self.font_small_bold = tkfont.Font(family="Microsoft YaHei UI", size=10, weight="bold")
            self.font_heading = tkfont.Font(family="Microsoft YaHei UI", size=13, weight="bold")
            self.font_title = tkfont.Font(family="Microsoft YaHei UI", size=18, weight="bold")
            self.font_hero = tkfont.Font(family="Microsoft YaHei UI", size=24, weight="bold")
            self.font_chat = tkfont.Font(family="Microsoft YaHei UI", size=12)
            self.font_input = tkfont.Font(family="Microsoft YaHei UI", size=12)
            self.option_add("*Font", self.font_ui)

            style = ttk.Style(self)
            try:
                style.theme_use("clam")
            except Exception:
                pass
            style.configure(".", font=self.font_ui, background=COLORS["app_bg"], foreground=COLORS["text"])
            style.configure("TCombobox", fieldbackground=COLORS["card_alt"], background=COLORS["card"], bordercolor=COLORS["line"], arrowcolor=COLORS["muted"])
            style.configure("TSpinbox", fieldbackground=COLORS["card_alt"], background=COLORS["card"], bordercolor=COLORS["line"], arrowcolor=COLORS["muted"])
            style.configure(
                "EnterpriseRetrieval.Horizontal.TProgressbar",
                thickness=6,
                troughcolor="#1F2A44",
                background=COLORS["primary"],
                lightcolor=COLORS["primary"],
                darkcolor=COLORS["primary"],
                bordercolor="#1F2A44",
            )

            self.configure(bg=COLORS["app_bg"])
            shell = tk.Frame(self, bg=COLORS["app_bg"])
            shell.pack(fill=tk.BOTH, expand=True)
            shell.columnconfigure(1, weight=1)
            shell.rowconfigure(0, weight=1)

            self.sidebar = Sidebar(shell, self)
            self.sidebar.grid(row=0, column=0, sticky=tk.NS)
            self.chat_area = ChatArea(shell, self)
            self.chat_area.grid(row=0, column=1, sticky=tk.NSEW)
            self.settings_panel = SettingsPanel(shell, self)
            self.settings_visible = False

            self.progress = ttk.Progressbar(
                self.sidebar.status_bar,
                variable=self.progress_var,
                maximum=100,
                mode="determinate",
                style="EnterpriseRetrieval.Horizontal.TProgressbar",
            )
            self.progress.pack(fill=tk.X, pady=(12, 0))
            self.sidebar.refresh()
            self.after_idle(self.render_chat_history)

        def toggle_settings_panel(self) -> None:
            visible = bool(getattr(self, "settings_visible", False))
            if visible:
                self.settings_panel.grid_remove()
                self.settings_visible = False
                if hasattr(self, "chat_area") and hasattr(self.chat_area, "settings_button"):
                    self.chat_area.settings_button.configure(text="模型与检索")
            else:
                self.settings_panel.grid(row=0, column=2, sticky=tk.NS)
                self.settings_visible = True
                if hasattr(self, "chat_area") and hasattr(self.chat_area, "settings_button"):
                    self.chat_area.settings_button.configure(text="收起设置")

        def choose_folder(self) -> None:
            folder = filedialog.askdirectory(title="选择要检索的文件夹")
            if folder:
                self.folder_var.set(folder)
                if hasattr(self, "sidebar"):
                    self.sidebar.refresh()
                self.status_var.set(f"已选择资料文件夹：{Path(folder).name}")

        def choose_db(self) -> None:
            path = filedialog.asksaveasfilename(
                title="选择索引库保存位置",
                defaultextension=".sqlite",
                filetypes=[("SQLite database", "*.sqlite"), ("All files", "*.*")],
            )
            if path:
                self.db_var.set(path)
                self.status_var.set(f"索引库：{Path(path).name}")

        def fill_example_question(self, question: str) -> None:
            if hasattr(self, "chat_input"):
                self.chat_input.delete("1.0", tk.END)
                self.chat_input.insert("1.0", question)
                self.chat_input.focus_set()

        def show_index_status(self) -> None:
            db_path = Path(self.db_var.get().strip())
            if not db_path.exists():
                self.render_chat_history(
                    pending_question="查看索引状态",
                    pending_answer="当前还没有可用索引。请先选择资料文件夹并建立索引。\n",
                )
                return
            try:
                conn = sqlite3.connect(str(db_path))
                files = conn.execute("SELECT COUNT(*) FROM files").fetchone()[0]
                chunks = conn.execute("SELECT COUNT(*) FROM chunks").fetchone()[0]
                conn.close()
                self.render_chat_history(
                    pending_question="查看索引状态",
                    pending_answer=f"当前索引库可用：已索引 {files} 个文件，生成 {chunks} 个文本片段。\n",
                )
            except Exception as exc:
                self.render_chat_history(
                    pending_question="查看索引状态",
                    pending_answer=f"读取索引状态失败：{exc}\n",
                    error=True,
                )

        def stop_rag_generation(self) -> None:
            if self.rag_worker and self.rag_worker.is_alive():
                self.rag_cancel_requested = True
                self.rag_status_var.set("已请求停止生成，当前请求返回后会忽略结果")
            else:
                self.rag_status_var.set("当前没有正在生成的回答")

        def copy_to_clipboard(self, text: str, label: str = "内容") -> None:
            text = str(text or "")
            if not text.strip():
                self.rag_status_var.set("没有可复制的内容")
                return
            try:
                self.clipboard_clear()
                self.clipboard_append(text)
                self.update_idletasks()
                self.rag_status_var.set(f"已复制{label}")
            except Exception as exc:
                self.rag_status_var.set(f"复制失败：{exc}")

        def compose_citation_text(self, citations: Optional[list[dict[str, str]]]) -> str:
            citations = citations or []
            lines: list[str] = []
            for item in citations:
                index = str(item.get("index", "")).strip()
                file_name = str(item.get("file", "未知文件")).strip()
                position = str(item.get("position", "")).strip()
                origin = str(item.get("origin", "本地检索")).strip()
                path = str(item.get("path", "")).strip()
                snippet = str(item.get("snippet", "")).strip()
                prefix = f"[{index}] " if index else ""
                block = [f"{prefix}{file_name}"]
                meta = " · ".join(part for part in (position, origin) if part)
                if meta:
                    block.append(meta)
                if path:
                    block.append(path)
                if snippet:
                    block.append(f"摘录：{snippet}")
                lines.append("\n".join(block))
            return "\n\n".join(lines)

        def compose_answer_with_citations(self, answer: str, citations: Optional[list[dict[str, str]]]) -> str:
            citation_text = self.compose_citation_text(citations)
            if citation_text:
                return f"{answer.strip()}\n\n引用资料：\n{citation_text}"
            return answer.strip()

        def open_citation(self, item: dict[str, str]) -> None:
            path = str(item.get("path", ""))
            if not path:
                return
            position = str(item.get("position", ""))
            page = 0
            page_match = re.search(r"PDF\s*第\s*(\d+)\s*页", position)
            if page_match:
                page = int(page_match.group(1))
            hit = SearchHit(
                path=path,
                keyword="",
                position=position,
                line=0,
                column=0,
                offset=0,
                chunk_no=0,
                snippet=str(item.get("snippet", "")),
                occurrence=1,
                page=page,
            )
            try:
                open_hit_target(hit)
            except Exception as exc:
                self.rag_status_var.set(f"打开引用失败：{exc}")

        def start_index(self) -> None:
            if self.worker and self.worker.is_alive():
                messagebox.showinfo("正在索引", "索引任务还在运行，请稍后。")
                return
            folder = Path(self.folder_var.get().strip())
            db_path = Path(self.db_var.get().strip())
            if not folder.exists() or not folder.is_dir():
                messagebox.showerror("文件夹无效", "请先选择有效的目标文件夹。")
                return
            try:
                if folder.anchor and folder.resolve() == Path(folder.anchor).resolve():
                    if not messagebox.askyesno("目录过大", "你选择的是磁盘根目录，可能导致电脑明显卡顿。确定继续吗？"):
                        return
            except Exception:
                pass

            self.stop_event.clear()
            self.progress_var.set(0)
            self.progress_percent_var.set("0%")
            self.progress_detail_var.set("准备索引...")
            self.status_var.set("正在建立索引...")
            self.index_button.configure(state=tk.DISABLED)
            self.stop_button.configure(state=tk.NORMAL)
            set_process_low_priority(True)

            def run() -> None:
                try:
                    stats = index_folder(
                        folder,
                        db_path,
                        max_file_mb=int(self.max_mb_var.get()),
                        progress=self.enqueue_progress,
                        should_stop=self.stop_event.is_set,
                    )
                    self.queue.put({"event": "index_cancelled" if stats.get("cancelled") else "index_success", **stats})
                except Exception as exc:
                    self.queue.put({"event": "index_error", "error": str(exc)})

            self.worker = threading.Thread(target=run, daemon=True)
            self.worker.start()

        def enqueue_progress(self, payload: dict[str, object]) -> None:
            try:
                self.queue.put_nowait(payload)
            except queue.Full:
                pass

        def stop_index(self) -> None:
            if self.worker and self.worker.is_alive():
                self.stop_event.set()
                self.stop_button.configure(state=tk.DISABLED)
                self.status_var.set("正在停止索引，请稍候...")
                self.progress_detail_var.set("正在停止...")

        def finish_index_ui(self) -> None:
            self.index_button.configure(state=tk.NORMAL)
            self.stop_button.configure(state=tk.DISABLED)
            set_process_low_priority(False)

        def update_index_progress(self, payload: dict[str, object]) -> None:
            total = int(payload.get("total") or 0)
            seen = int(payload.get("seen") or 0)
            if total > 0:
                percent = min(100.0, max(0.0, seen * 100.0 / total))
                self.progress_var.set(percent)
                self.progress_percent_var.set(f"{percent:.0f}%")
                self.progress_detail_var.set(
                    f"已处理 {seen}/{total}，索引 {payload.get('indexed', 0)}，"
                    f"跳过 {payload.get('skipped', 0)}，失败 {payload.get('failed', 0)}"
                )
            path = str(payload.get("path", ""))
            if path:
                self.status_var.set(f"正在索引：{Path(path).name}")

        def drain_queue(self) -> None:
            try:
                while True:
                    payload = self.queue.get_nowait()
                    event = payload.get("event")
                    if event == "counting":
                        self.progress_var.set(0)
                        self.progress_percent_var.set("0%")
                        self.progress_detail_var.set(f"正在统计文件：已发现 {payload.get('total', 0)} 个")
                    elif event == "counted":
                        self.progress_detail_var.set(f"共 {payload.get('total', 0)} 个文件，开始索引")
                    elif event == "file":
                        self.update_index_progress(payload)
                    elif event == "index_success":
                        self.progress_var.set(100)
                        self.progress_percent_var.set("100%")
                        self.progress_detail_var.set(
                            f"完成：索引 {payload.get('indexed', 0)}，跳过 {payload.get('skipped', 0)}，失败 {payload.get('failed', 0)}"
                        )
                        self.status_var.set(
                            "索引完成："
                            f"扫描 {payload.get('seen', 0)}，"
                            f"索引 {payload.get('indexed', 0)}，"
                            f"跳过 {payload.get('skipped', 0)}，"
                            f"失败 {payload.get('failed', 0)}"
                        )
                        self.finish_index_ui()
                    elif event in {"index_cancelled", "cancelled"}:
                        self.progress_detail_var.set(
                            f"已停止：处理 {payload.get('seen', 0)}/{payload.get('total', 0)}"
                        )
                        self.status_var.set("索引已停止")
                        self.finish_index_ui()
                    elif event == "index_error":
                        self.finish_index_ui()
                        messagebox.showerror("索引失败", str(payload.get("error", "")))
                        self.status_var.set("索引失败")
                    elif event == "rag_success":
                        self.finish_rag_ui()
                        question = str(payload.get("question", ""))
                        answer = str(payload.get("answer", ""))
                        sources = payload.get("sources", [])
                        source_items = payload.get("source_items", [])
                        source_text = "\n".join(str(item) for item in sources) if isinstance(sources, list) else ""
                        if not isinstance(source_items, list):
                            source_items = []
                        self.add_rag_history(question, answer, source_text, source_items=source_items)
                        self.rag_status_var.set("回答完成")
                    elif event == "rag_progress":
                        question = str(payload.get("question", ""))
                        message = str(payload.get("message", ""))
                        self.rag_status_var.set(message)
                        if question:
                            self.render_chat_history(pending_question=question, pending_answer=message + "\n")
                    elif event == "rag_cancelled":
                        self.finish_rag_ui()
                        question = str(payload.get("question", ""))
                        self.rag_status_var.set("已停止生成")
                        if question:
                            self.render_chat_history(pending_question=question, pending_answer="已停止生成，本次回答未写入历史。\n")
                    elif event == "rag_error":
                        self.finish_rag_ui()
                        error = str(payload.get("error", ""))
                        question = str(payload.get("question", ""))
                        self.write_error_log(error)
                        self.rag_status_var.set("回答失败")
                        if question:
                            self.render_chat_history(pending_question=question, pending_answer=f"问答失败：{error}\n", error=True)
                        else:
                            self.set_rag_answer(error)
            except queue.Empty:
                pass
            self.after(150, self.drain_queue)

        def save_current_rag_settings(self) -> None:
            try:
                save_rag_settings(
                    self.rag_api_url_var.get(),
                    self.rag_model_var.get(),
                    self.rag_api_key_var.get(),
                    max(1, int(self.rag_top_k_var.get())),
                    self.rag_save_key_var.get(),
                )
                self.rag_status_var.set("配置已保存")
            except Exception as exc:
                messagebox.showerror("保存失败", str(exc))

        def write_error_log(self, message: str) -> None:
            try:
                log_path = default_db_path().with_name("rag_error.log")
                log_path.parent.mkdir(parents=True, exist_ok=True)
                with log_path.open("a", encoding="utf-8") as f:
                    f.write(f"[{now_iso()}] {message}\n")
            except Exception:
                pass

        def set_rag_answer(self, text: str) -> None:
            if hasattr(self, "chat_area"):
                self.chat_area.hide_welcome()
                self.chat_area.add_message("assistant", text, error=True, scroll="top")
                return

        def clean_chat_markdown(self, text: str) -> str:
            text = normalize_text(text)
            cleaned_lines: list[str] = []
            for raw_line in text.splitlines():
                line = raw_line.strip()
                if not line:
                    cleaned_lines.append("")
                    continue
                line = re.sub(r"^\s*[\*\-]\s+", "• ", line)
                line = re.sub(r"^•\s+\*\*(.+?)\*\*\s*[:：]", r"• \1：", line)
                line = re.sub(r"\*\*(.+?)\*\*", r"\1", line)
                line = re.sub(r"`([^`]+)`", r"\1", line)
                cleaned_lines.append(line)
            text = "\n".join(cleaned_lines)
            text = re.sub(r"\n{3,}", "\n\n", text)
            return text.strip()

        def format_sources_for_chat(self, sources: str) -> str:
            lines = [line.strip() for line in sources.splitlines() if line.strip()]
            formatted: list[str] = []
            max_visible = 8
            for line in lines[:max_visible]:
                line = line.replace(" | ", "  /  ")
                formatted.append(line)
            if len(lines) > max_visible:
                formatted.append(f"还有 {len(lines) - max_visible} 条引用已用于回答")
            return "\n".join(formatted)

        def append_chat_block(self, role: str, text: str, turn_no: int = 0, sources: str = "") -> None:
            if role == "user":
                title = f"你 · 第 {turn_no} 轮" if turn_no else "你"
                self.rag_answer_text.insert(tk.END, title + "\n", ("user_label",))
                self.rag_answer_text.insert(tk.END, text.strip() + "\n\n", ("user_message",))
                return

            self.rag_answer_text.insert(tk.END, DISPLAY_NAME + "\n", ("assistant_label",))
            display_text = self.clean_chat_markdown(text)
            self.rag_answer_text.insert(tk.END, display_text + "\n", ("assistant_message",))
            if sources:
                self.rag_answer_text.insert(tk.END, "\n资料引用\n", ("source_label",))
                self.rag_answer_text.insert(tk.END, self.format_sources_for_chat(sources) + "\n", ("source",))
            self.rag_answer_text.insert(tk.END, "\n", ("message_gap",))

        def render_chat_history(self, pending_question: str = "", pending_answer: str = "", error: bool = False) -> None:
            if hasattr(self, "chat_area"):
                self.chat_area.render(
                    self.rag_history,
                    pending_question=pending_question,
                    pending_answer=pending_answer,
                    error=error,
                )

        def format_rag_history(self) -> str:
            if not self.rag_history:
                return ""
            parts: list[str] = []
            turn_no = 1
            for index in range(0, len(self.rag_history), 2):
                user_item = self.rag_history[index]
                assistant_item = self.rag_history[index + 1] if index + 1 < len(self.rag_history) else {}
                question = user_item.get("content", "")
                answer = assistant_item.get("content", "")
                sources = assistant_item.get("sources", "")
                block = [f"第 {turn_no} 轮", f"问：{question}", "", f"答：{answer}"]
                if sources:
                    block.extend(["", "来源：", sources])
                parts.append("\n".join(block))
                turn_no += 1
            return ("\n\n" + "-" * 60 + "\n\n").join(parts)

        def add_rag_history(
            self,
            question: str,
            answer: str,
            sources: str,
            source_items: Optional[list[dict[str, str]]] = None,
        ) -> None:
            if not question:
                question = self.query_var.get().strip()
            self.rag_history.append({"role": "user", "content": question})
            self.rag_history.append(
                {
                    "role": "assistant",
                    "content": answer,
                    "sources": sources,
                    "source_items": source_items or [],
                }
            )
            trimmed = False
            if len(self.rag_history) > 16:
                self.rag_history = self.rag_history[-16:]
                trimmed = True
            if trimmed and hasattr(self, "chat_area"):
                self.chat_area.clear_messages()
            self.render_chat_history()

        def clear_rag_history(self) -> None:
            self.rag_history.clear()
            self.last_hits = []
            self.render_chat_history()
            self.rag_status_var.set("对话已清空")

        def finish_rag_ui(self) -> None:
            self.rag_button.configure(state=tk.NORMAL)
            if hasattr(self, "chat_input"):
                self.chat_input.focus_set()

        def current_question_text(self) -> str:
            if hasattr(self, "chat_input"):
                return self.chat_input.get("1.0", tk.END).strip()
            return self.query_var.get().strip()

        def clear_question_text(self) -> None:
            if hasattr(self, "chat_input"):
                self.chat_input.delete("1.0", tk.END)
            self.query_var.set("")

        def handle_chat_return(self, event: object) -> Optional[str]:
            state = int(getattr(event, "state", 0))
            if state & 0x0001:
                return None
            self.run_rag_answer()
            return "break"

        def run_rag_answer(self) -> None:
            if self.rag_worker and self.rag_worker.is_alive():
                messagebox.showinfo("正在生成", "当前回答还在生成，请稍后。")
                return
            question = self.current_question_text()
            if not question:
                messagebox.showinfo("请输入问题", "请在下方聊天框里输入要询问的问题。")
                return
            db_path = Path(self.db_var.get().strip())
            self.query_var.set(question)

            api_url = self.rag_api_url_var.get().strip()
            model = self.rag_model_var.get().strip()
            api_key = self.rag_api_key_var.get().strip()
            if not api_key and "localhost" not in api_url and "127.0.0.1" not in api_url:
                self.rag_status_var.set("缺少 API Key")
                self.render_chat_history(
                    pending_question=question,
                    pending_answer="API Key 未填写。请在右侧“模型设置”中填写 API Key 后再发送问题。\n",
                    error=True,
                )
                return
            top_k = max(1, min(20, int(self.rag_top_k_var.get())))
            match_case = self.match_case_var.get()
            require_all = self.require_all_var.get()
            phrase = self.phrase_var.get()
            history_snapshot = [dict(item) for item in self.rag_history]
            context_only = should_skip_local_retrieval(question, history_snapshot)
            if not context_only and not db_path.exists():
                self.rag_status_var.set("缺少索引")
                self.render_chat_history(
                    pending_question=question,
                    pending_answer="请先选择资料文件夹并建立索引，然后再提问。\n",
                    error=True,
                )
                return
            try:
                save_rag_settings(api_url, model, api_key, top_k, self.rag_save_key_var.get())
            except Exception:
                pass

            self.rag_button.configure(state=tk.DISABLED)
            self.rag_cancel_requested = False
            self.clear_question_text()
            initial_message = "正在结合上下文生成回答…" if context_only else "正在检索本地知识库…"
            self.rag_status_var.set(initial_message)
            self.render_chat_history(pending_question=question, pending_answer=initial_message + "\n")

            def run() -> None:
                try:
                    if context_only:
                        self.queue.put(
                            {"event": "rag_progress", "question": question, "message": "正在结合上下文生成回答…"}
                        )
                        answer = call_context_completion(
                            api_url,
                            api_key,
                            model,
                            question,
                            chat_history=history_snapshot,
                        )
                        if self.rag_cancel_requested:
                            self.queue.put({"event": "rag_cancelled", "question": question})
                            return
                        self.queue.put(
                            {
                                "event": "rag_success",
                                "question": question,
                                "answer": answer,
                                "sources": [],
                                "source_items": [],
                            }
                        )
                        return
                    self.queue.put({"event": "rag_progress", "question": question, "message": "正在检索本地知识库…"})
                    retrieval_query = retrieval_query_for_question(question, history_snapshot)
                    file_candidates = rank_candidate_files(
                        db_path,
                        retrieval_query,
                        match_case=match_case,
                        limit=max(top_k * 8, 32),
                    )
                    preferred_file_paths = {item.path for item in file_candidates}
                    if preferred_file_paths:
                        self.queue.put(
                            {
                                "event": "rag_progress",
                                "question": question,
                                "message": f"已按文件名和内容线索筛选 {len(preferred_file_paths)} 个候选文件，正在检索内容…",
                            }
                        )
                    priority_hits: list[SearchHit] = []
                    seen_hits: set[tuple[str, str, int]] = set()
                    for candidate in rag_search_candidates(retrieval_query):
                        candidate_hits = search_index(
                            db_path,
                            candidate,
                            require_all_terms=require_all,
                            match_case=match_case,
                            phrase=phrase,
                            limit=max(top_k * 4, top_k),
                            file_paths=preferred_file_paths if preferred_file_paths else None,
                        )
                        for hit in candidate_hits:
                            key = (hit.path, hit.keyword.casefold() if not match_case else hit.keyword, hit.offset)
                            if key in seen_hits:
                                continue
                            seen_hits.add(key)
                            priority_hits.append(hit)
                            if len(priority_hits) >= max(top_k * 4, top_k):
                                break
                        if len(priority_hits) >= max(top_k * 8, top_k):
                            break
                    if not priority_hits and preferred_file_paths:
                        self.queue.put(
                            {
                                "event": "rag_progress",
                                "question": question,
                                "message": "候选文件内未找到精确命中，正在全库补充检索…",
                            }
                        )
                        for candidate in rag_search_candidates(retrieval_query):
                            candidate_hits = search_index(
                                db_path,
                                candidate,
                                require_all_terms=require_all,
                                match_case=match_case,
                                phrase=phrase,
                                limit=max(top_k * 4, top_k),
                            )
                            for hit in candidate_hits:
                                key = (hit.path, hit.keyword.casefold() if not match_case else hit.keyword, hit.offset)
                                if key in seen_hits:
                                    continue
                                seen_hits.add(key)
                                priority_hits.append(hit)
                                if len(priority_hits) >= max(top_k * 4, top_k):
                                    break
                            if len(priority_hits) >= max(top_k * 8, top_k):
                                break
                    if self.rag_cancel_requested:
                        self.queue.put({"event": "rag_cancelled", "question": question})
                        return
                    contexts = fetch_rag_contexts(
                        db_path,
                        question,
                        require_all_terms=require_all,
                        match_case=match_case,
                        phrase=phrase,
                        limit=top_k,
                        chat_history=history_snapshot,
                        priority_hits=priority_hits,
                        preferred_file_paths=preferred_file_paths,
                    )
                    self.queue.put(
                        {
                            "event": "rag_progress",
                            "question": question,
                            "message": f"已找到 {len(contexts)} 条相关资料，正在生成回答…",
                        }
                    )
                    if self.rag_cancel_requested:
                        self.queue.put({"event": "rag_cancelled", "question": question})
                        return
                    answer = call_chat_completion(
                        api_url,
                        api_key,
                        model,
                        question,
                        contexts,
                        chat_history=history_snapshot,
                    )
                    if self.rag_cancel_requested:
                        self.queue.put({"event": "rag_cancelled", "question": question})
                        return
                    sources = [
                        f"[{index}] {Path(context.path).name} | {context.position} | {context.origin or '本地检索'}"
                        for index, context in enumerate(contexts, start=1)
                    ]
                    source_items = []
                    for index, context in enumerate(contexts, start=1):
                        focus_terms = unique_keep_order([context.keyword, *context.matched_terms, *rag_question_terms(question)])
                        source_items.append(
                            {
                                "index": str(index),
                                "file": Path(context.path).name,
                                "path": context.path,
                                "position": context.position,
                                "origin": context.origin or "本地检索",
                                "snippet": focused_context_text(context.text, focus_terms, 220),
                            }
                        )
                    self.queue.put(
                        {
                            "event": "rag_success",
                            "question": question,
                            "answer": answer,
                            "sources": sources,
                            "source_items": source_items,
                        }
                    )
                except Exception as exc:
                    self.queue.put({"event": "rag_error", "question": question, "error": str(exc)})

            self.rag_worker = threading.Thread(target=run, daemon=True)
            self.rag_worker.start()

        def run_search(self) -> None:
            query = self.query_var.get().strip()
            if not query:
                messagebox.showinfo("请输入关键词", "请输入一个或多个关键词。包含空格的词组可用英文引号包起来。")
                return
            try:
                hits = search_index(
                    Path(self.db_var.get().strip()),
                    query,
                    require_all_terms=self.require_all_var.get(),
                    match_case=self.match_case_var.get(),
                    phrase=self.phrase_var.get(),
                    limit=int(self.limit_var.get()),
                )
            except Exception as exc:
                messagebox.showerror("搜索失败", str(exc))
                return
            self.last_hits = hits
            self.status_var.set(f"找到 {len(hits)} 条命中")
            if not hasattr(self, "tree"):
                self.rag_status_var.set(f"本地预检索找到 {len(hits)} 条命中")
                return
            self.tree.delete(*self.tree.get_children())
            for index, hit in enumerate(hits):
                self.tree.insert(
                    "",
                    tk.END,
                    iid=str(index),
                    tags=("hit",),
                    values=(
                        hit.path,
                        hit.keyword,
                        hit.position,
                    ),
                )
            if hits:
                self.tree.selection_set("0")
                self.tree.focus("0")
                self.update_context_preview()
            else:
                self.clear_context_preview()

        def selected_hit(self) -> Optional[SearchHit]:
            if not hasattr(self, "tree"):
                return None
            selected = self.tree.selection()
            if not selected:
                return None
            try:
                return self.last_hits[int(selected[0])]
            except (ValueError, IndexError):
                return None

        def clear_context_preview(self) -> None:
            if not hasattr(self, "context_text"):
                return
            self.context_text.configure(state=tk.NORMAL)
            self.context_text.delete("1.0", tk.END)
            self.context_text.configure(state=tk.DISABLED)

        def update_context_preview(self) -> None:
            if not hasattr(self, "context_text"):
                return
            hit = self.selected_hit()
            self.context_text.configure(state=tk.NORMAL)
            self.context_text.delete("1.0", tk.END)
            if hit:
                self.context_text.insert(tk.END, hit.snippet)
                flags = 0 if self.match_case_var.get() else re.IGNORECASE
                for match in re.finditer(re.escape(hit.keyword), hit.snippet, flags=flags):
                    start = f"1.0+{match.start()}c"
                    end = f"1.0+{match.end()}c"
                    self.context_text.tag_add("keyword", start, end)
            self.context_text.configure(state=tk.DISABLED)

        def open_selected_file(self) -> None:
            hit = self.selected_hit()
            if not hit:
                return
            try:
                open_hit_target(hit)
            except Exception as exc:
                messagebox.showerror("打开失败", str(exc))

        def reveal_selected_file(self) -> None:
            hit = self.selected_hit()
            if not hit:
                return
            try:
                import subprocess

                subprocess.Popen(["explorer", "/select,", hit.path])
            except Exception as exc:
                messagebox.showerror("打开失败", str(exc))

        def export_csv(self) -> None:
            if not self.last_hits:
                messagebox.showinfo("没有结果", "当前没有可导出的搜索结果。")
                return
            path = filedialog.asksaveasfilename(
                title="导出搜索结果",
                defaultextension=".csv",
                filetypes=[("CSV", "*.csv"), ("All files", "*.*")],
            )
            if path:
                export_hits_csv(Path(path), self.last_hits)
                self.status_var.set(f"已导出：{path}")

    App().mainloop()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=f"{DISPLAY_NAME}: index a Windows folder and search keywords.")
    sub = parser.add_subparsers(dest="command")

    sub.add_parser("gui", help="open desktop GUI")

    index_parser = sub.add_parser("index", help="build or rebuild an index")
    index_parser.add_argument("folder", help="folder to index")
    index_parser.add_argument("--db", default=str(default_db_path()), help="SQLite index path")
    index_parser.add_argument("--chunk-size", type=int, default=DEFAULT_CHUNK_SIZE)
    index_parser.add_argument("--overlap", type=int, default=DEFAULT_CHUNK_OVERLAP)
    index_parser.add_argument("--max-file-mb", type=int, default=DEFAULT_MAX_FILE_MB)
    index_parser.add_argument("--max-text-chars", type=int, default=DEFAULT_MAX_TEXT_CHARS)

    search_parser = sub.add_parser("search", help="search an existing index")
    search_parser.add_argument("query", help='keyword query, e.g. "invoice customer" or "\\"full phrase\\""')
    search_parser.add_argument("--db", default=str(default_db_path()), help="SQLite index path")
    search_parser.add_argument("--all", action="store_true", help="require all keywords in the same chunk")
    search_parser.add_argument("--match-case", action="store_true", help="case-sensitive match")
    search_parser.add_argument("--phrase", action="store_true", help="treat the whole query as one exact phrase")
    search_parser.add_argument("--limit", type=int, default=500)
    search_parser.add_argument("--json", action="store_true", help="print JSON instead of text table")
    search_parser.add_argument("--csv", help="write CSV result file")

    return parser


def main(argv: Optional[list[str]] = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)

    if args.command in {None, "gui"}:
        launch_gui()
        return 0

    if args.command == "index":
        def progress(payload: dict[str, object]) -> None:
            if payload.get("event") == "file":
                seen = payload.get("seen", 0)
                if int(seen) % 50 == 0:
                    print(
                        f"scanned={payload.get('seen', 0)} "
                        f"indexed={payload.get('indexed', 0)} "
                        f"skipped={payload.get('skipped', 0)} "
                        f"failed={payload.get('failed', 0)}"
                    )

        stats = index_folder(
            Path(args.folder),
            Path(args.db),
            chunk_size=args.chunk_size,
            overlap=args.overlap,
            max_file_mb=args.max_file_mb,
            max_text_chars=args.max_text_chars,
            progress=progress,
        )
        print(json.dumps(stats, ensure_ascii=False, indent=2))
        print(f"index: {Path(args.db).resolve()}")
        return 0

    if args.command == "search":
        hits = search_index(
            Path(args.db),
            args.query,
            require_all_terms=args.all,
            match_case=args.match_case,
            phrase=args.phrase,
            limit=args.limit,
        )
        if args.csv:
            export_hits_csv(Path(args.csv), hits)
            print(f"wrote: {Path(args.csv).resolve()}")
        elif args.json:
            print(json.dumps([hit.__dict__ for hit in hits], ensure_ascii=False, indent=2))
        else:
            print_hits_table(hits)
        return 0

    parser.print_help()
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
